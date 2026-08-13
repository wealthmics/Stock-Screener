r"""
=====================================================================================
MARQUEE INVESTOR PORTFOLIO  +  GICS SECTOR SCREENER
=====================================================================================
One script. Put the two CSV files in a folder, run it, and you get the formatted
workbook plus an interactive HTML screener. When new data arrives, replace the CSV
files and run again. No code changes needed.

-------------------------------------------------------------------------------------
STEP 1 - WHERE TO PUT THE CSV FILES
-------------------------------------------------------------------------------------
  Create a folder, put BOTH CSV files in it, and write that folder path into DATA_DIR
  below. The folder can be named anything.

      screener/
        marquee_screener.py                  <- this script
        data/                                <- DATA_DIR   (put the CSV files here)
          <any name>.csv                         1) All Stocks screener export
          <any name>.csv                         2) Dataroma marquee export
        output/                              <- OUT_DIR    (created automatically)
          Marquee_Investor_Portfolio_Consolidated.xlsx
          Sector_Screener.html

  FILENAMES DO NOT MATTER. The script opens every CSV in the folder and works out
  which is which from the column headers, so you can rename or re-export the files
  as often as you like. The All Stocks export is recognised by columns such as
  "Market capitalization" and "Price to earnings ratio"; the Dataroma export by its
  ownership columns such as "Ownership count" and "Hold Price".
  If an older copy of either export is still sitting in the folder, the most recently
  modified one is used and the older one is reported and skipped. Any CSV that matches
  neither export is ignored.

  The Dataroma half can also be scraped for you instead of exported by hand - see
  SCRAPE_DATAROMA below. In that case only the All Stocks CSV needs to be in the folder.

-------------------------------------------------------------------------------------
STEP 2 - RUN IT
-------------------------------------------------------------------------------------
  Once DATA_DIR is set:

      python3 marquee_screener.py

  Or pass the folders on the command line instead:

      python3 marquee_screener.py --data-dir "D:/work/screener/data" --out-dir "D:/work/screener/output"

  In Jupyter, Spyder or a VS Code notebook, run it as a cell like this:

      %run marquee_screener.py

  The notebook is detected automatically and its own arguments are ignored, so the
  DATA_DIR and OUT_DIR values below are what get used. Windows paths need the r prefix:

      DATA_DIR = r'C:\Users\LENOVO\Desktop\Stock Research\CSV'
      OUT_DIR  = r'C:\Users\LENOVO\Desktop\Stock Research\CSV\OUTPUT'

-------------------------------------------------------------------------------------
WHAT TO CHANGE AND WHERE  (seven places, no need to read the rest of the code)
-------------------------------------------------------------------------------------
  [0] DATA_DIR / OUT_DIR   Where the CSV files live and where output goes. Top of file.
  [1] SCREEN_UNIVERSE      'all' screens the full 30,000 stock universe.
                           'marquee' screens only the superinvestor holdings.
  [2] DEFAULT_FILTERS      The base threshold set. Since every sector now has its own
                           complete block, this only applies to a sector you add to
                           SECTOR_MAP without giving it a block. All four margins are
                           here - gross, operating, net and free cash flow - and the
                           FCF margin filter is switched on.
  [3] SECTOR_OVERRIDES     THE MAIN ONE. Every GICS sector has its own block, and each
                           block lists all fifteen metrics explicitly. Nothing is
                           hidden and nothing arrives silently from the defaults.
                           Change a value directly, or write None to switch that
                           filter off. Banks never clear ROA above 10% and utilities
                           never clear debt to equity below 2, so a few sectors carry
                           values that differ from the defaults - those lines are
                           marked "# <- changed from default".
                           Nothing passing? Open that sector's block, loosen a value
                           and run again.
                           'All Sectors' is the first block. It screens the whole
                           universe in one go and feeds the "Screen - All Sectors"
                           sheet, and it is editable exactly like the rest.
  [4] BLANK_FAILS /        True rejects any stock missing a screened value.
      MIN_DATA_COVERAGE    False ignores the blank and judges the stock on what it
                           does report. Altman Z covers about 76% of the universe and
                           Piotroski about 63%, so True returns far fewer names. When
                           BLANK_FAILS is False, MIN_DATA_COVERAGE is what stops a
                           thinly covered stock from passing by default.
  [5] SECTOR_MAP           The GICS name against its TradingView sector list. Your
                           mapping is already in place. Add a new sector here.

  You do not need to open Excel to check the thresholds. The "Screen Settings" sheet
  lists the final threshold set for every sector, with overrides flagged in a
  different colour.

-------------------------------------------------------------------------------------
THE INTERACTIVE HTML
-------------------------------------------------------------------------------------
  Sector_Screener.html lands in the output folder. Double-click it in a browser. No
  server and no internet connection needed, because the data is embedded in the file.
  There you can change every threshold live, switch sector (including All Sectors),
  toggle US against non-US, and download the result as CSV.
  To push the screener to a live website on every run, switch GITHUB_UPLOAD on and set
  a token. Setup steps are written above that setting.

  Changes made in the HTML are temporary. Once you settle on a threshold set, use the
  "Copy as Python" button and paste the block into SECTOR_OVERRIDES below. Only then
  does it become permanent in the workbook.
=====================================================================================
"""

import argparse, datetime, glob, json, os, re, sys, time
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =====================================================================================
# [0] FOLDERS  and  [1] UNIVERSE
# =====================================================================================
# Write your own folder here. The script uses these, so nothing needs to be typed on
# the command line.
#   Windows:      DATA_DIR = r'C:\Users\Shobhit\Desktop\screener\data'
#   Mac / Linux:  DATA_DIR = '/Users/shobhit/Desktop/screener/data'
DATA_DIR = 'data'
OUT_DIR  = 'output'

SCREEN_UNIVERSE = 'all'    # 'all' = full universe, 'marquee' = superinvestor holdings only

# --------------------------------------------------------------------------------------
# Dataroma scraping. The marquee holdings can be pulled straight off dataroma.com with
# Selenium instead of exporting a CSV by hand.
#   'never'   use the dataroma CSV already sitting in DATA_DIR
#   'auto'    scrape only when there is no dataroma CSV, or the newest one is older than
#             DATAROMA_MAX_AGE_DAYS. This is the sensible default.
#   'always'  scrape on every run
# Needs Chrome plus:  pip install selenium
# The scraped file is written into DATA_DIR as dataroma_portfolio_YYYY-MM-DD.csv, so the
# normal auto-detection picks it up like any other export.
SCRAPE_DATAROMA = 'auto'
DATAROMA_MAX_AGE_DAYS = 3
DATAROMA_PAGES = 15        # the scraper stops early on its own if a page comes back empty

# Which threshold set to run.
#   'base'   every sector uses the identical DEFAULT_FILTERS values (SECTOR_OVERRIDES).
#            This is the starting point - one rule for the whole market.
#   'tuned'  uses SECTOR_PRESETS instead. Those thresholds are calibrated against the
#            actual marquee holdings in each sector - each bar sits near that sector's
#            median - and ROIC and ROCE are switched off everywhere as redundant with
#            ROE and ROA. This is what produces a workable count in every sector.
# Switch this one word and run again. Both sets are fully editable further down.
THRESHOLD_MODE = 'tuned'

# How the "Screen - All Sectors" sheet is built.
#   True   every row is judged against ITS OWN sector's block, and the results are then
#          combined into one list. A bank is judged on the Financials rules, a REIT on the
#          Real Estate rules, and so on. The combined list is exactly the union of the
#          eleven sector sheets. Rows whose TradingView sector maps to no GICS sector
#          (Miscellaneous, Government) are judged on the 'All Sectors' block instead.
#   False  one single rule set - the 'All Sectors' block - is applied to all 30,000 rows.
#          Stricter and like-for-like, but it wipes out sectors that structurally cannot
#          clear a market-wide rule.
ALL_SECTORS_USES_SECTOR_RULES = True

# Put the data date into the output filenames, e.g.
# Marquee_Investor_Portfolio_Consolidated_2026-08-10.xlsx. The date is read from the CSV
# filename, and from the file's modified date if the name carries no date.
# Set False for stable names, which is easier if you re-upload the HTML to a website.
DATE_IN_FILENAME = True

# Optional: a local folder that is also a git repository for a website, typically a
# GitHub Pages repo. When this is set, the screener is copied there as index.html on
# every run, so publishing is just: run this script, then git add / commit / push.
# Leave it as '' to switch the feature off.
#   PUBLISH_DIR = r'C:\Users\LENOVO\Desktop\mics-screener'
PUBLISH_DIR = ''

# --------------------------------------------------------------------------------------
# Deep analysis app. When this is set, every ticker in the HTML screener becomes a link
# to a Streamlit app that pulls live data for that one company and runs the forensic
# scores. Leave it as '' and the tickers stay plain text.
# Streamlit is the right home for that page because the fetch happens server side -
# a static page cannot call Yahoo directly, since Yahoo sends no CORS header.
#   DEEP_APP_URL = 'https://your-app.streamlit.app'
DEEP_APP_URL = 'https://mics-deep-analysis-dcxigpgnhvhywvq2599eh3.streamlit.app'

# --------------------------------------------------------------------------------------
# GitHub upload. With this on, the screener is pushed straight to the repo at the end of
# every run, so the live site updates without opening a browser.
#
# One-time setup:
#   1. github.com -> click your avatar -> Settings (your account, not the repo)
#   2. Developer settings -> Personal access tokens -> Fine-grained tokens
#   3. Generate new token. Repository access: only your screener repo.
#      Permissions -> Repository permissions -> Contents: Read and write. Nothing else.
#   4. Copy the token (starts with github_pat_) and set it as an environment variable
#      called GITHUB_TOKEN, or paste it below.
#
# Windows, to set it permanently, run once in Command Prompt then reopen your editor:
#      setx GITHUB_TOKEN github_pat_xxxxxxxxxxxxxxxx
#
# A token is a password. Never share it, never screenshot it, and never let this script
# sit in a public repo with a real token pasted in. If one ever leaks, revoke it on the
# same GitHub page and generate a new one.
GITHUB_UPLOAD = False  # the workflow commits instead
GITHUB_REPO   = 'wealthmics/Stock-Screener'   # owner/repo
GITHUB_BRANCH = 'main'
GITHUB_PATH   = 'index.html'                  # the name GitHub Pages serves
GITHUB_XLSX   = 'Marquee_Investor_Portfolio_Consolidated.xlsx'   # '' to skip the workbook
# the deep analysis app fetches these two from the Pages site, so they have to be committed
GITHUB_FEED   = ['stocks.json', 'sector_stats.json']
# The token is written in below as you asked. Because it sits inside this file, do not
# upload this .py to the public Stock-Screener repo - GitHub's secret scanner will find it
# and disable the token. To move it out later, set the line back to '' and run once in
# Command Prompt:   setx GITHUB_TOKEN github_pat_your_token_here   then restart Jupyter.
# When the line is empty the script reads the GITHUB_TOKEN environment variable instead.
GITHUB_TOKEN  = ''

# How the HTML is written.
#   False  one self-contained file, about 6 MB. Double-click and it works, no internet,
#          no server. Best for emailing the file or keeping it on your own machine.
#   True   a light index.html plus a separate data.json. The page appears instantly and
#          the data loads behind a short "loading" message. This is what a website wants,
#          including GitHub Pages. It will NOT work by double-clicking from your desktop,
#          because browsers block local file reads.
# Both files are written to OUT_DIR, and to PUBLISH_DIR as index.html + data.json when set.
WEB_SPLIT = False

# =====================================================================================
# [4] MISSING DATA POLICY
# =====================================================================================
BLANK_FAILS = False        # True = strict, a blank value rejects the stock

# With BLANK_FAILS = False, a row that reports nothing at all would technically clear every
# filter. This guard stops that: a stock must report at least this share of the active
# filters to qualify. 0.6 means 60%. Set it to 0 to switch the guard off.
MIN_DATA_COVERAGE = 0.6

# =====================================================================================
# [2] DEFAULT FILTERS  ---  the base rule set
#     None = filter off.  '_min' means at least this value, '_max' means at most this value.
# =====================================================================================
DEFAULT_FILTERS = {
    'pe_max'            : 20,      # Price to earnings (TTM), at most
    'roe_min'           : 10,      # Return on equity %
    'roic_min'          : 10,      # Return on invested capital %
    'roce_min'          : 10,      # Return on capital employed %
    'roa_min'           : 10,      # Return on assets %
    'gross_margin_min'  : 10,      # Gross margin %
    'op_margin_min'     : 10,      # Operating margin %
    'net_margin_min'    : 10,      # Net margin %
    'fcf_margin_min'    : 10,      # Free cash flow margin %
    'altman_min'        : 2.99,    # Altman Z, above 2.99 is the safe zone
    'piotroski_min'     : 7,       # Piotroski F score out of 9
    'current_ratio_min' : 1,       # Current ratio
    'de_max'            : 2,       # Debt to equity, at most
    'int_cov_min'       : 3,       # Interest coverage
    'mcap_min'          : None,    # put a number here to drop micro caps
}

# =====================================================================================
# [3] SECTOR THRESHOLDS  ---  two complete sets, switched by THRESHOLD_MODE above
# =====================================================================================
#     Every block below lists all fifteen metrics explicitly. Nothing is hidden and
#     nothing arrives silently from somewhere else. Change a value directly, or write
#     None to switch that filter off.
#
#     SECTOR_OVERRIDES  is the 'base' set. Every sector starts on exactly the same
#                       thresholds you specified, so the screen is like-for-like across
#                       the market. This is what runs by default.
#     SECTOR_PRESETS    is the 'tuned' set, and the one that runs by default. Every bar is
#                       calibrated against the marquee holdings actually in that sector,
#                       anchored near the sector median, with the comment above each block
#                       quoting those medians. ROIC and ROCE are off everywhere because
#                       they correlate above 80% with ROE and ROA - running all four at
#                       once collapses the joint pass rate without adding information.
#                       Lines that differ from the base carry '# <- differs from the base set'.
#
#     Switch between them with THRESHOLD_MODE at the top. Edit either one freely.
# =====================================================================================
SECTOR_OVERRIDES = {

    # ---- All Sectors -----------------------------------------------------------
    # The whole universe in one pass, sector split ignored.
    'All Sectors': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Information Technology ------------------------------------------------
    'Information Technology': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Communication Services ------------------------------------------------
    'Communication Services': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Consumer Discretionary ------------------------------------------------
    'Consumer Discretionary': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Consumer Staples ------------------------------------------------------
    'Consumer Staples': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Health Care -----------------------------------------------------------
    'Health Care': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Financials ------------------------------------------------------------
    'Financials': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Energy ----------------------------------------------------------------
    'Energy': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Materials -------------------------------------------------------------
    'Materials': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Industrials -----------------------------------------------------------
    'Industrials': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Utilities -------------------------------------------------------------
    'Utilities': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Real Estate -----------------------------------------------------------
    'Real Estate': {
        'pe_max'            : 20    ,
        'roe_min'           : 10    ,
        'roic_min'          : 10    ,
        'roce_min'          : 10    ,
        'roa_min'           : 10    ,
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 2.99  ,
        'piotroski_min'     : 7     ,
        'current_ratio_min' : 1     ,
        'de_max'            : 2     ,
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

}


# ---- the optional sector-tuned set, used when THRESHOLD_MODE = 'tuned' --------------
SECTOR_PRESETS = {

    # ---- All Sectors -----------------------------------------------------------
    # One relaxed set for the whole market. Sits roughly at the cross-sector median of
    # the marquee holdings, so the strict base rules do not wipe out entire sectors.
    'All Sectors': {
        'pe_max'            : 25    ,   # <- differs from the base set
        'roe_min'           : 10    ,
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 3     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 5     ,   # <- differs from the base set
        'fcf_margin_min'    : 5     ,   # <- differs from the base set
        'altman_min'        : 2     ,   # <- differs from the base set
        'piotroski_min'     : 5     ,   # <- differs from the base set
        'current_ratio_min' : 1     ,
        'de_max'            : 1.5   ,   # <- differs from the base set
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Information Technology ------------------------------------------------
    # Marquee medians here: P/E 39.6, ROE 12.8, ROA 6.1, Altman 4.5. A P/E cap of 20
    # only admits 40% of the sector and ROA of 10% only 35%, so both are relaxed.
    'Information Technology': {
        'pe_max'            : 45    ,   # <- differs from the base set
        'roe_min'           : 12    ,   # <- differs from the base set
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 5     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 10    ,
        'net_margin_min'    : 7     ,   # <- differs from the base set
        'fcf_margin_min'    : 10    ,
        'altman_min'        : 3     ,   # <- differs from the base set
        'piotroski_min'     : 6     ,   # <- differs from the base set
        'current_ratio_min' : 1.2   ,   # <- differs from the base set
        'de_max'            : 1     ,   # <- differs from the base set
        'int_cov_min'       : 4     ,   # <- differs from the base set
        'mcap_min'          : None  ,
    },

    # ---- Communication Services ------------------------------------------------
    # The thinnest sector - only 18 marquee names and weak fundamentals
    # (median ROE 7.1, Altman 1.5, interest coverage 0.9). Everything is set low deliberately;
    # tightening anything here empties the sheet.
    'Communication Services': {
        'pe_max'            : 25    ,   # <- differs from the base set
        'roe_min'           : 5     ,   # <- differs from the base set
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 1.5   ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 6     ,   # <- differs from the base set
        'net_margin_min'    : 3     ,   # <- differs from the base set
        'fcf_margin_min'    : 3     ,   # <- differs from the base set
        'altman_min'        : 1     ,   # <- differs from the base set
        'piotroski_min'     : 4     ,   # <- differs from the base set
        'current_ratio_min' : 0.8   ,   # <- differs from the base set
        'de_max'            : 2     ,
        'int_cov_min'       : 0.8   ,   # <- differs from the base set
        'mcap_min'          : None  ,
    },

    # ---- Consumer Discretionary ------------------------------------------------
    # Marquee medians: P/E 21.6, ROE 12.1, operating margin 8.0, net margin 4.4.
    # Volume business, so the margin bars sit below the base 10%.
    'Consumer Discretionary': {
        'pe_max'            : 28    ,   # <- differs from the base set
        'roe_min'           : 10    ,
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 3     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 7     ,   # <- differs from the base set
        'net_margin_min'    : 3.5   ,   # <- differs from the base set
        'fcf_margin_min'    : 4     ,   # <- differs from the base set
        'altman_min'        : 2.2   ,   # <- differs from the base set
        'piotroski_min'     : 6     ,   # <- differs from the base set
        'current_ratio_min' : 0.9   ,   # <- differs from the base set
        'de_max'            : 1.5   ,   # <- differs from the base set
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Consumer Staples ------------------------------------------------------
    # Marquee medians: P/E 23.0, ROE 12.2, net margin 4.5. Same thin-margin logic
    # as retail - net margin at 10% removes almost the whole sector.
    'Consumer Staples': {
        'pe_max'            : 28    ,   # <- differs from the base set
        'roe_min'           : 10    ,
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 3     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 8     ,   # <- differs from the base set
        'net_margin_min'    : 3     ,   # <- differs from the base set
        'fcf_margin_min'    : 4     ,   # <- differs from the base set
        'altman_min'        : 2     ,   # <- differs from the base set
        'piotroski_min'     : 6     ,   # <- differs from the base set
        'current_ratio_min' : 1     ,
        'de_max'            : 1.5   ,   # <- differs from the base set
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Health Care -----------------------------------------------------------
    # Marquee medians: P/E 25.4, ROE 5.5, gross margin 60.3. Pharma carries heavy R&D
    # so ROE and ROA run low while gross margin runs very high.
    'Health Care': {
        'pe_max'            : 30    ,   # <- differs from the base set
        'roe_min'           : 8     ,   # <- differs from the base set
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 3     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 12    ,   # <- differs from the base set
        'net_margin_min'    : 6     ,   # <- differs from the base set
        'fcf_margin_min'    : 8     ,   # <- differs from the base set
        'altman_min'        : 2.5   ,   # <- differs from the base set
        'piotroski_min'     : 5     ,   # <- differs from the base set
        'current_ratio_min' : 1.5   ,   # <- differs from the base set
        'de_max'            : 1     ,   # <- differs from the base set
        'int_cov_min'       : 4     ,   # <- differs from the base set
        'mcap_min'          : None  ,
    },

    # ---- Financials ------------------------------------------------------------
    # Banks: Altman Z, current ratio, debt to equity, interest coverage, gross margin and
    # FCF margin are all meaningless here and stay off. Median ROA is 1.5%, not 10%. Median
    # P/E is 14.2, so the cheapness bar can be tighter than the base rule.
    'Financials': {
        'pe_max'            : 16    ,   # <- differs from the base set
        'roe_min'           : 13    ,   # <- differs from the base set
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 1.5   ,   # <- differs from the base set
        'gross_margin_min'  : None  ,   # <- differs from the base set
        'op_margin_min'     : 25    ,   # <- differs from the base set
        'net_margin_min'    : 15    ,   # <- differs from the base set
        'fcf_margin_min'    : None  ,   # <- differs from the base set
        'altman_min'        : None  ,   # <- differs from the base set
        'piotroski_min'     : 5     ,   # <- differs from the base set
        'current_ratio_min' : None  ,   # <- differs from the base set
        'de_max'            : None  ,   # <- differs from the base set
        'int_cov_min'       : None  ,   # <- differs from the base set
        'mcap_min'          : None  ,
    },

    # ---- Energy ----------------------------------------------------------------
    # Marquee medians: P/E 10.6, ROE 13.9, ROA 6.9, operating margin 16.2. Cyclical, so
    # the P/E bar tightens while the margin bars loosen.
    'Energy': {
        'pe_max'            : 15    ,   # <- differs from the base set
        'roe_min'           : 10    ,
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 4     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 12    ,   # <- differs from the base set
        'net_margin_min'    : 6     ,   # <- differs from the base set
        'fcf_margin_min'    : 5     ,   # <- differs from the base set
        'altman_min'        : 1.8   ,   # <- differs from the base set
        'piotroski_min'     : 6     ,   # <- differs from the base set
        'current_ratio_min' : 0.9   ,   # <- differs from the base set
        'de_max'            : 1     ,   # <- differs from the base set
        'int_cov_min'       : 3     ,
        'mcap_min'          : None  ,
    },

    # ---- Materials -------------------------------------------------------------
    # Marquee medians: P/E 21.7, ROE 9.2, ROA 4.4. Capital heavy commodity business.
    'Materials': {
        'pe_max'            : 25    ,   # <- differs from the base set
        'roe_min'           : 9     ,   # <- differs from the base set
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 4     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 12    ,   # <- differs from the base set
        'net_margin_min'    : 5     ,   # <- differs from the base set
        'fcf_margin_min'    : 6     ,   # <- differs from the base set
        'altman_min'        : 2.5   ,   # <- differs from the base set
        'piotroski_min'     : 6     ,   # <- differs from the base set
        'current_ratio_min' : 1.5   ,   # <- differs from the base set
        'de_max'            : 1     ,   # <- differs from the base set
        'int_cov_min'       : 4     ,   # <- differs from the base set
        'mcap_min'          : None  ,
    },

    # ---- Industrials -----------------------------------------------------------
    # Marquee medians: P/E 27.8, ROE 13.7, ROA 4.9, operating margin 12.4.
    'Industrials': {
        'pe_max'            : 30    ,   # <- differs from the base set
        'roe_min'           : 13    ,   # <- differs from the base set
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 4     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 12    ,   # <- differs from the base set
        'net_margin_min'    : 7     ,   # <- differs from the base set
        'fcf_margin_min'    : 8     ,   # <- differs from the base set
        'altman_min'        : 2.5   ,   # <- differs from the base set
        'piotroski_min'     : 6     ,   # <- differs from the base set
        'current_ratio_min' : 1.4   ,   # <- differs from the base set
        'de_max'            : 1     ,   # <- differs from the base set
        'int_cov_min'       : 4     ,   # <- differs from the base set
        'mcap_min'          : None  ,
    },

    # ---- Utilities -------------------------------------------------------------
    # Only 25 marquee names. Median FCF margin is negative (-7.4) because of capex and
    # median Altman is 1.1, so both are off or set very low. Median operating margin is 21.9,
    # which lets that bar sit higher than the base rule.
    'Utilities': {
        'pe_max'            : 28    ,   # <- differs from the base set
        'roe_min'           : 8     ,   # <- differs from the base set
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 2     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 15    ,   # <- differs from the base set
        'net_margin_min'    : 8     ,   # <- differs from the base set
        'fcf_margin_min'    : None  ,   # <- differs from the base set
        'altman_min'        : 0.8   ,   # <- differs from the base set
        'piotroski_min'     : 5     ,   # <- differs from the base set
        'current_ratio_min' : None  ,   # <- differs from the base set
        'de_max'            : 3     ,   # <- differs from the base set
        'int_cov_min'       : 1.5   ,   # <- differs from the base set
        'mcap_min'          : None  ,
    },

    # ---- Real Estate -----------------------------------------------------------
    # REITs: median Altman 1.4 and current ratio 1.1, so those stay off. Median operating
    # margin 19.3 and net margin 13.2 are high, so those bars go above the base rule.
    'Real Estate': {
        'pe_max'            : 32    ,   # <- differs from the base set
        'roe_min'           : 7     ,   # <- differs from the base set
        'roic_min'          : None  ,   # <- differs from the base set
        'roce_min'          : None  ,   # <- differs from the base set
        'roa_min'           : 3     ,   # <- differs from the base set
        'gross_margin_min'  : 10    ,
        'op_margin_min'     : 18    ,   # <- differs from the base set
        'net_margin_min'    : 10    ,
        'fcf_margin_min'    : None  ,   # <- differs from the base set
        'altman_min'        : None  ,   # <- differs from the base set
        'piotroski_min'     : 5     ,   # <- differs from the base set
        'current_ratio_min' : None  ,   # <- differs from the base set
        'de_max'            : 2     ,
        'int_cov_min'       : 1.5   ,   # <- differs from the base set
        'mcap_min'          : None  ,
    },

}


# =====================================================================================
# [5] GICS  <-  TradingView SECTOR MAPPING
# =====================================================================================
SECTOR_MAP = {
    'Information Technology' : ['Electronic technology', 'Technology services'],
    'Communication Services' : ['Communications'],
    'Consumer Discretionary' : ['Consumer durables', 'Consumer services', 'Retail trade'],
    'Consumer Staples'       : ['Consumer non-durables', 'Distribution services'],
    'Health Care'            : ['Health technology', 'Health services'],
    'Financials'             : ['Finance'],
    'Energy'                 : ['Energy minerals'],
    'Materials'              : ['Non-energy minerals', 'Process industries'],
    'Industrials'            : ['Producer manufacturing', 'Industrial services',
                                'Commercial services', 'Transportation'],
    'Utilities'              : ['Utilities'],
    'Real Estate'            : [],   # no separate TradingView sector - see the note below
}

# TradingView has no separate Real Estate sector, so REITs mostly sit inside Finance.
# This pattern pulls them out by industry name onto the Real Estate sheet and removes
# them from the Financials sheet.
REAL_ESTATE_INDUSTRY = re.compile(r'real estate|reit', re.I)
PULL_REITS_OUT_OF_FINANCE = True     # set False to leave the Real Estate sheet empty

ALL_LABEL = 'All Sectors'      # the combined, no-sector-split screen

# Excel sheet names cannot exceed 31 characters
SHEET_NAMES = {
    ALL_LABEL                : 'Screen - All Sectors',
    'Information Technology' : 'Screen - Info Technology',
    'Communication Services' : 'Screen - Communication Svcs',
    'Consumer Discretionary' : 'Screen - Cons Discretionary',
    'Consumer Staples'       : 'Screen - Cons Staples',
    'Health Care'            : 'Screen - Health Care',
    'Financials'             : 'Screen - Financials',
    'Energy'                 : 'Screen - Energy',
    'Materials'              : 'Screen - Materials',
    'Industrials'            : 'Screen - Industrials',
    'Utilities'              : 'Screen - Utilities',
    'Real Estate'            : 'Screen - Real Estate',
}

# =====================================================================================
# ==============  ENGINE BELOW THIS LINE - normally nothing to change here  ============
# =====================================================================================

# filter key -> (source column, direction, pretty label)
FILTER_SPECS = {
    'pe_max'            : ('Price to earnings ratio',                          'max', 'P/E (TTM)'),
    'roe_min'           : ('Return on equity %, Trailing 12 months',           'min', 'ROE %'),
    'roic_min'          : ('Return on invested capital %, Trailing 12 months',  'min', 'ROIC %'),
    'roce_min'          : ('Return on capital employed %, Trailing 12 months',  'min', 'ROCE %'),
    'roa_min'           : ('Return on assets %, Trailing 12 months',            'min', 'ROA %'),
    'gross_margin_min'  : ('Gross margin %, Annual',                            'min', 'Gross margin %'),
    'op_margin_min'     : ('Operating margin %, Annual',                        'min', 'Operating margin %'),
    'net_margin_min'    : ('Net margin %, Annual',                              'min', 'Net margin %'),
    'fcf_margin_min'    : ('Free cash flow margin %, Annual',                   'min', 'FCF margin %'),
    'altman_min'        : ('Altman Z-score, Trailing 12 months',                'min', 'Altman Z'),
    'piotroski_min'     : ('Piotroski F-score, Trailing 12 months',             'min', 'Piotroski F'),
    'current_ratio_min' : ('Current ratio, Quarterly',                          'min', 'Current ratio'),
    'de_max'            : ('Debt to equity ratio, Quarterly',                   'max', 'Debt / Equity'),
    'int_cov_min'       : ('Interest coverage, Annual',                         'min', 'Interest coverage'),
    'mcap_min'          : ('Market capitalization',                             'min', 'Market cap'),
}

PCT, NUM, X, BIG, TXT = '0.0"%";(0.0"%");-', '#,##0.00', '0.0"x"', '#,##0', '@'
Z, F = '0.00', '0"/9"'
USD = '$#,##0.00'

thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
H1F = Font(name='Arial', size=10, bold=True, color='FFFFFF')
H2F = Font(name='Arial', size=9, bold=True, color='FFFFFF')
BF = Font(name='Arial', size=9)
BOLD = Font(name='Arial', size=9, bold=True)
CEN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right')
STRIPE = PatternFill('solid', fgColor='F2F5FA')

NAVY, GOLD, PLUM = '1F3864', 'BF8F00', '4B2E83'


# ------------------------------------------------------------------ data loading -----
DATE_RE = re.compile(r'(20\d{2})[-_. ]?(\d{2})[-_. ]?(\d{2})')


def file_date(path):
    """The data date, taken from the filename if it carries one, else the file's own date."""
    m = DATE_RE.search(os.path.basename(path))
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return datetime.date.fromtimestamp(os.path.getmtime(path))


def pretty_date(d):
    return f'{d.day} {d:%B %Y}'


def _header(path):
    """Column names of a CSV, lower-cased, without reading the whole file."""
    try:
        return [str(c).strip().lower().replace('\n', ' ')
                for c in pd.read_csv(path, nrows=0).columns]
    except Exception:
        return []


def classify(path):
    """Work out what a CSV is from its columns, not its filename.

    Returns 'universe', 'marquee' or None. This is why the files can be named
    anything - rename them freely, re-export them, keep the date suffix or drop it.
    """
    cols = _header(path)
    if not cols:
        return None
    joined = ' | '.join(cols)
    # The All Stocks export carries the screener's financial columns.
    universe_hits = sum(k in joined for k in
                        ('market capitalization', 'price to earnings', 'return on equity',
                         'net margin', 'country or region', 'sector'))
    # The Dataroma export is a short ownership table: ticker, name, weight, holder count.
    marquee_hits = sum(k in joined for k in
                       ('ownership', 'hold price', 'portfolio', '52 week', '52w', 'max %'))
    has_pct_col = any(c.strip().startswith('%') or c.strip('  ') in ('%', '%\u25bc') for c in cols)
    if universe_hits >= 3:
        return 'universe'
    if (marquee_hits >= 2 or (has_pct_col and 'stock' in joined)) and len(cols) <= 15:
        return 'marquee'
    return None


def find_files(data_dir):
    """Auto-detect the universe and marquee CSVs in a folder by their contents.

    Filenames are never used to decide what a file is. If several files of the same
    kind are present, the most recently modified one wins, so an old export left in
    the folder is harmless.
    """
    if not os.path.isdir(data_dir):
        sys.exit(f"\n[ERROR] The folder does not exist:\n        {os.path.abspath(data_dir)}\n"
                 f"        Fix: correct DATA_DIR at the top of the script.\n")
    csvs = sorted(glob.glob(os.path.join(data_dir, '*.csv')))
    found = {'universe': [], 'marquee': [], 'unknown': []}
    for path in csvs:
        found[classify(path) or 'unknown'].append(path)

    def newest(paths):
        return max(paths, key=os.path.getmtime) if paths else None

    uni, mrq = newest(found['universe']), newest(found['marquee'])
    if uni and mrq:
        for kind, chosen in (('universe', uni), ('marquee', mrq)):
            extra = [os.path.basename(x) for x in found[kind] if x != chosen]
            if extra:
                print(f'      note: {len(extra)} older {kind} file(s) ignored -> {extra}')
        if found['unknown']:
            print(f"      note: ignored {[os.path.basename(x) for x in found['unknown']]} "
                  f"(columns match neither export)")
        return uni, mrq

    listing = '\n'.join(
        f'          {os.path.basename(p):<45} detected as: {classify(p) or "unrecognised"}'
        for p in csvs) or '          no CSV files at all'
    missing = 'All Stocks screener export' if not uni else 'Dataroma marquee export'
    sys.exit(f"\n[ERROR] Could not find the {missing} in this folder.\n"
             f"        Looked in : {os.path.abspath(data_dir)}\n"
             f"        Saw       :\n{listing}\n"
             f"        The two files are identified by their columns, not their names.\n"
             f"        The All Stocks export needs columns such as 'Market capitalization' and\n"
             f"        'Price to earnings ratio'. The Dataroma export needs its ownership\n"
             f"        columns such as 'Ownership count' and 'Hold Price'.\n"
             f"        Fix       : re-export the missing file, or check DATA_DIR.\n")


def to_money(s):
    if pd.isna(s):
        return np.nan
    return pd.to_numeric(str(s).replace('$', '').replace(',', '').strip(), errors='coerce')


SUFFIX = (r'\b(incorporated|inc|corporation|corp|company|co|limited|ltd|plc|llc|lp|nv|n\.v|sa|s\.a'
          r'|ag|ab|asa|as|a/s|oyj|spa|s\.p\.a|holdings?|holding|group|grp|the|trust|reit|adr|ads'
          r'|sponsored|cl|class|series)\b')


def norm_name(n):
    s = str(n).lower().replace('&', ' and ')
    s = re.sub(r'[^a-z0-9/\. ]', ' ', s)
    s = re.sub(r'\.', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    for _ in range(4):
        s = re.sub(SUFFIX, ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
    for _ in range(2):
        s = re.sub(r'\b(a|b|c|k|h|i|ii)\b$', '', s).strip()
    return re.sub(r'\s+', ' ', s).strip()


# Hand-checked ADR / alternate share class map: dataroma ticker -> (universe ticker, country)
ALIAS = {
    'TSM': ('2330', 'Taiwan'), 'NVO': ('NOVO_B', 'Denmark'), 'NSRGY': ('NESN', 'Switzerland'),
    'DB': ('DBK', 'Germany'), 'SNY': ('SAN', 'France'), 'FMS': ('FME', 'Germany'),
    'LZAGY': ('LONN', 'Switzerland'), 'NGG': ('NG.', 'United Kingdom'),
    'RYCEY': ('RR.', 'United Kingdom'), 'LOGI': ('LOGN', 'Switzerland'),
    'FLUIF': ('FDR', 'Spain'), 'HMRZF': ('HM_B', 'Sweden'), 'PHJMF': ('HMSP', 'Indonesia'),
    'DVDCF': ('CPR', None), 'BUD': ('ABI', 'Belgium'), 'BUDFF': ('ABI', 'Belgium'),
    'ABEV': ('ABEV3', 'Brazil'), 'ITUB': ('ITUB3', 'Brazil'), 'VALE': ('VALE3', 'Brazil'),
    'EMBJ': ('EMBJ3', 'Brazil'), 'YPF': ('YPFD', 'Argentina'), 'TEO': ('TECO2', 'Argentina'),
    'CIB': ('CIBEST', 'Colombia'), 'FMX': ('FEMSA/UB', 'Mexico'), 'ASR': ('ASUR/B', 'Mexico'),
    'ZTO': ('2057', 'China'), 'SMFG': ('8316', 'Japan'), 'PBR.A': ('PETR3', 'Brazil'),
    'UHAL.B': ('UHAL', 'United States'), 'FWONK': ('FWONA', 'United States'),
    'LLYVK': ('LLYVA', 'United States'), 'GLIBK': ('GLIBA', 'United States'),
    'GLBAV-OLD': ('GLIBA', 'United States'), 'HTZWW': ('HTZ', 'United States'),
    'AUROW': ('AUR', 'United States'), 'ASPSW': ('ASPS', 'Luxembourg'),
    'ASPSZ': ('ASPS', 'Luxembourg'), 'ILLUU': ('ILLU', 'United States'),
}

ETF_TICKERS = set(
    'SPY VOO IWM GLD SGOV EWY FRDM QQQ DIA IVV VTI AGG TLT HYBB BIL SHV IEF LQD SLV IAU VEA VWO '
    'EFA IEMG JEPI SCHD VIG DVY XLF XLE XLK XLV XLI XLP XLU XLB XLY XLC VNQ IJH IJR MDY RSP ITOT '
    'SPLG VXUS BND VCIT VCSH JNK USHY EMB PFF ARKK SMH SOXX IBIT FBTC GBTC ETHE MOAT KWEB GDXJ '
    'DINT MAVF AMLP IWF'.split())
ETF_HINT = re.compile(r'\b(etf|ishares|spdr|vanguard|index|invesco|proshares|select sector|treasury'
                      r'|gold shares|fund|msci|s&p 500|russell|kraneshares|vaneck|alerian'
                      r'|davis select)\b', re.I)


ISIN_COL_HINT = re.compile(r'isin', re.I)


def find_isin_column(df):
    """The ISIN column, whatever TradingView decided to call it, or None."""
    for c in df.columns:
        if ISIN_COL_HINT.search(str(c)):
            return c
    return None


def load_universe(path):
    a = pd.read_csv(path)
    for col in ('Altman Z-score, Trailing 12 months', 'Piotroski F-score, Trailing 12 months'):
        if col not in a.columns:
            a[col] = np.nan
            print(f"[WARN] '{col}' is not in the CSV. Any filter on that column is ignored.")
    a['_sym'] = a['Symbol'].astype(str).str.strip()
    a['_norm'] = a['Description'].map(norm_name)
    a['_us'] = (a['Country or region of registration'] == 'United States').astype(int)
    isin_col = find_isin_column(a)
    if isin_col:
        a['ISIN'] = a[isin_col].astype(str).str.strip().str.upper()
        good = a['ISIN'].str.fullmatch(r'[A-Z]{2}[A-Z0-9]{9}[0-9]', na=False)
        a.loc[~good, 'ISIN'] = None
        print(f"      ISIN column '{isin_col}' found - {int(good.sum())} of {len(a)} rows "
              f"carry a valid ISIN")
    else:
        a['ISIN'] = None
        print('      no ISIN column in this export - deep links will pass the ticker only')
    return a


def load_marquee(path):
    d = pd.read_csv(path)
    d.columns = ['Symbol', 'Stock', 'Portfolio %', 'Ownership Count', 'Hold Price', 'Max %',
                 'Current Price', '52W Low', '% Above 52W Low', '52W High']
    for c in ('Hold Price', 'Current Price', '52W Low', '52W High'):
        d[c] = d[c].map(to_money)
    d = d.drop_duplicates(subset=['Symbol', 'Stock']).reset_index(drop=True)
    d['Symbol'] = d['Symbol'].astype(str).str.strip()
    d['Stock'] = d['Stock'].astype(str).str.strip()
    d['_norm'] = d['Stock'].map(norm_name)
    return d



# ------------------------------------------------------------------ dataroma scrape ---
def scrape_dataroma(data_dir):
    """Pull the superinvestor aggregate holdings off dataroma.com into a CSV.

    Selenium is imported here rather than at the top of the file so the script still runs
    for anyone who only has the CSV exports and never scrapes.
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait
    except ImportError:
        sys.exit('\n[ERROR] Selenium is not installed, so Dataroma cannot be scraped.\n'
                 '        Fix: pip install selenium     (Chrome must also be installed)\n'
                 "        Or set SCRAPE_DATAROMA = 'never' and put the CSV in DATA_DIR "
                 'by hand.\n')

    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1400,1000')
    options.add_experimental_option('prefs', {
        'profile.managed_default_content_settings.images': 2,
        'profile.default_content_setting_values.notifications': 2,
    })

    url = 'https://www.dataroma.com/m/g/portfolio.php?L={}'
    headers, rows_out = [], []
    driver = None
    try:
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 20)
        for page in range(1, DATAROMA_PAGES + 1):
            got = 0
            for attempt in (1, 2, 3):
                try:
                    driver.get(url.format(page))
                    table = wait.until(EC.presence_of_element_located((By.ID, 'grid')))
                    trs = table.find_elements(By.TAG_NAME, 'tr')
                    if not trs:
                        break
                    if not headers:
                        headers = [td.text.strip()
                                   for td in trs[0].find_elements(By.TAG_NAME, 'td')]
                    for tr in trs[1:]:
                        cols = [td.text.strip() for td in tr.find_elements(By.TAG_NAME, 'td')]
                        if cols and len(cols) == len(headers):
                            rows_out.append(cols)
                            got += 1
                    break
                except Exception as e:
                    if attempt == 3:
                        print(f'      page {page} failed after 3 tries, skipped ({e.__class__.__name__})')
                    continue
            print(f'      page {page:>2} -> {got} holdings')
            if got == 0 and page > 1:
                print('      empty page, stopping early')
                break
    finally:
        if driver is not None:
            driver.quit()

    if not headers or len(rows_out) < 100:
        sys.exit(f'\n[ERROR] The Dataroma scrape returned only {len(rows_out)} rows, which '
                 f'means the page layout changed or the site blocked the request.\n'
                 f"        Fix: set SCRAPE_DATAROMA = 'never' and export the CSV by hand "
                 f'for now.\n')

    df = pd.DataFrame(rows_out, columns=headers).drop_duplicates()
    out = os.path.join(data_dir, f'dataroma_portfolio_{datetime.date.today():%Y-%m-%d}.csv')
    df.to_csv(out, index=False)
    print(f'      scraped {len(df)} unique holdings -> {os.path.basename(out)}')
    return out


def dataroma_needs_scrape(data_dir):
    """Decide whether to scrape, based on SCRAPE_DATAROMA and how old the local file is."""
    if SCRAPE_DATAROMA == 'always':
        return True, 'SCRAPE_DATAROMA is always'
    if SCRAPE_DATAROMA == 'never':
        return False, ''
    existing = [p for p in glob.glob(os.path.join(data_dir, '*.csv'))
                if classify(p) == 'marquee']
    if not existing:
        return True, 'no dataroma CSV in the folder'
    newest = max(existing, key=os.path.getmtime)
    age = (datetime.date.today() - file_date(newest)).days
    if age > DATAROMA_MAX_AGE_DAYS:
        return True, f'{os.path.basename(newest)} is {age} days old'
    return False, f'{os.path.basename(newest)} is {age} days old, fresh enough'



# ------------------------------------------------------------------ matching ---------
def match_marquee(a, d):
    by_sym, by_norm = a.groupby('_sym').groups, a.groupby('_norm').groups

    def pick(cands, px, country=None):
        if len(cands) == 0:
            return None
        c = cands
        if country is not None:
            cc = c[c['Country or region of registration'] == country]
            if len(cc):
                c = cc
        if len(c) == 1:
            return c.index[0]
        if not pd.isna(px):
            t = c.copy()
            t['_pd'] = (pd.to_numeric(t['Price'], errors='coerce') - px).abs() / px
            close = t[t['_pd'] < 0.02]
            if len(close):
                return close.sort_values('_pd').index[0]
        if c['_us'].max() == 1:
            c = c[c['_us'] == 1]
        return c.sort_values('Market capitalization', ascending=False).index[0]

    recs = []
    for _, r in d.iterrows():
        sym, px = r['Symbol'], r['Current Price']
        idx, method = None, 'No match'
        if sym in ALIAS:
            tgt, ctry = ALIAS[sym]
            if tgt in by_sym:
                idx = pick(a.loc[by_sym[tgt]], px, ctry)
                method = 'Verified cross-listing / share-class alias'
        if idx is None:
            for cs, tag in [(sym, 'Exact ticker'),
                            (re.sub(r'-OLD$', '', sym), 'Ticker (legacy -OLD stripped)'),
                            (sym.replace('.', '_'), 'Ticker variant'),
                            (sym.replace('.', '-'), 'Ticker variant')]:
                if cs in by_sym:
                    j = pick(a.loc[by_sym[cs]], px)
                    if j is not None:
                        idx, method = j, tag
                        break
        if idx is None and r['_norm'] and r['_norm'] in by_norm:
            j = pick(a.loc[by_norm[r['_norm']]], px)
            if j is not None:
                idx, method = j, 'Company name match'

        rec = dict(r)
        rec['Match Method'] = method
        if idx is not None:
            m = a.loc[idx]
            rec['Matched Symbol'], rec['Matched Company'] = m['Symbol'], m['Description']
            for c in a.columns:
                if not c.startswith('_'):
                    rec['A_' + c] = m[c]
            p = pd.to_numeric(m['Price'], errors='coerce')
            same_px = (not pd.isna(px)) and (not pd.isna(p)) and abs(p - px) / px < 0.03
            alt_class = (m['Symbol'] != sym) and (not pd.isna(px)) and (not pd.isna(p)) \
                        and abs(p - px) / px > 0.20
            if method == 'Exact ticker' and same_px:
                rec['Match Confidence'] = 'High'
            elif method == 'Exact ticker':
                rec['Match Confidence'] = 'High' if m['Price - Currency'] == 'USD' else 'Medium'
            elif method in ('Verified cross-listing / share-class alias', 'Company name match'):
                rec['Match Confidence'] = 'High' if same_px else 'Medium'
            else:
                rec['Match Confidence'] = 'Medium'
            if m['Price - Currency'] == 'USD' and alt_class:
                rec['Data Basis'] = (f"Alternate share class ({m['Symbol']}) - company ratios "
                                     f"identical, per-share figures relate to {m['Symbol']}")
            elif m['Price - Currency'] == 'USD':
                rec['Data Basis'] = ('US line item' if (pd.isna(px) or pd.isna(p)
                                     or abs(p - px) / px <= 0.20)
                                     else 'US line item - screener price differs from the marquee '
                                          'list quote, verify before use')
            else:
                rec['Data Basis'] = (f"Primary listing in {m['Country or region of registration']} "
                                     f"({m['Price - Currency']}) - ratios comparable, price and FCF "
                                     f"in local currency")
        else:
            rec['Matched Symbol'] = None
            if sym in ETF_TICKERS or ETF_HINT.search(r['Stock']):
                rec['Reason Not Identified'] = 'ETF / index fund / trust - no company-level financials exist'
            elif sym.endswith('-OLD'):
                rec['Reason Not Identified'] = 'Legacy or superseded ticker (-OLD) - not in the current equity universe'
            elif re.search(r'(WTS|W$|U$|Z$)', sym) or 'WTS' in r['Stock'].upper():
                rec['Reason Not Identified'] = 'Warrant / unit / non-common security - no underlying match found'
            else:
                rec['Reason Not Identified'] = 'Company not present in the All Stocks universe (ADR-only, OTC or delisted)'
        recs.append(rec)
    return pd.DataFrame(recs)


# ------------------------------------------------------------------ screening --------
def screen_order():
    """All Sectors first, then the eleven GICS sectors."""
    return [ALL_LABEL] + list(SECTOR_MAP)


def active_set():
    """The threshold dictionary THRESHOLD_MODE points at."""
    if THRESHOLD_MODE not in ('base', 'tuned'):
        sys.exit(f"[ERROR] THRESHOLD_MODE must be 'base' or 'tuned', not {THRESHOLD_MODE!r}.")
    return SECTOR_PRESETS if THRESHOLD_MODE == 'tuned' else SECTOR_OVERRIDES


def sector_filters(gics):
    """DEFAULT_FILTERS plus that sector's block from the active set = the final rules."""
    f = dict(DEFAULT_FILTERS)
    f.update(active_set().get(gics, {}))
    return f


def assign_gics(df):
    tv2gics = {}
    for g, tvs in SECTOR_MAP.items():
        for tv in tvs:
            tv2gics[tv.strip().lower()] = g
    g = df['Sector'].astype(str).str.strip().str.lower().map(tv2gics)
    if PULL_REITS_OUT_OF_FINANCE:
        is_re = df['Industry'].astype(str).str.contains(REAL_ESTATE_INDUSTRY)
        g = g.where(~is_re, 'Real Estate')
    return g


def screen(df, gics):
    f = sector_filters(gics)
    mask = pd.Series(True, index=df.index)
    reported = pd.Series(0, index=df.index)
    applied, n_active = [], 0
    for key, val in f.items():
        if val is None:
            continue
        col, direction, label = FILTER_SPECS[key]
        series = pd.to_numeric(df[col], errors='coerce')
        ok = (series >= val) if direction == 'min' else (series <= val)
        if not BLANK_FAILS:
            ok = ok | series.isna()
        mask &= ok.fillna(False)
        reported += series.notna().astype(int)
        n_active += 1
        applied.append((label, ('>=' if direction == 'min' else '<='), val))
    if not BLANK_FAILS and n_active and MIN_DATA_COVERAGE > 0:
        mask &= (reported / n_active) >= MIN_DATA_COVERAGE
    return df[mask].copy(), applied, f


# ------------------------------------------------------------------ excel helpers ----
def write_grid(ws, df, groups, freeze_col=3, start_row=1):
    """Two-row banded header + striped data rows. Returns (ncols, {src: col_index})."""
    col, fmts = 1, []
    hr1, hr2 = start_row, start_row + 1
    for gname, gcolor, cols in groups:
        start = col
        for label, src, fmt in cols:
            c = ws.cell(row=hr2, column=col, value=label)
            c.font, c.fill = H2F, PatternFill('solid', fgColor=gcolor)
            c.alignment, c.border = CEN, BORDER
            fmts.append((col, src, fmt))
            col += 1
        g = ws.cell(row=hr1, column=start, value=gname)
        g.font, g.fill = H1F, PatternFill('solid', fgColor=gcolor)
        g.alignment, g.border = CEN, BORDER
        if col - 1 > start:
            ws.merge_cells(start_row=hr1, start_column=start, end_row=hr1, end_column=col - 1)
        for cc in range(start, col):
            ws.cell(row=hr1, column=cc).fill = PatternFill('solid', fgColor=gcolor)
            ws.cell(row=hr1, column=cc).border = BORDER
    ncols = col - 1
    for i, (_, row) in enumerate(df.iterrows()):
        r = hr2 + 1 + i
        band = (i % 2 == 1)
        for cidx, src, fmt in fmts:
            c = ws.cell(row=r, column=cidx)
            if not src.startswith('__'):
                v = row.get(src, None)
                if isinstance(v, float) and pd.isna(v):
                    v = None
                if isinstance(v, str) and v.strip() == '':
                    v = None
                c.value = v
            c.font, c.number_format, c.border = BF, fmt, BORDER
            c.alignment = LEFT if fmt == TXT else RIGHT
            if band:
                c.fill = STRIPE
    ws.freeze_panes = ws.cell(row=hr2 + 1, column=freeze_col)
    ws.auto_filter.ref = f"A{hr2}:{get_column_letter(ncols)}{hr2 + len(df)}"
    ws.row_dimensions[hr1].height = 20
    ws.row_dimensions[hr2].height = 34
    for cidx, src, fmt in fmts:
        hdr = str(ws.cell(row=hr2, column=cidx).value)
        w = max(9, min(30, len(hdr) + 2))
        if fmt == TXT:
            w = max(w, 14)
        if src in ('Stock', 'A_Description', 'Description', 'A_Industry', 'Industry',
                   'Data Basis', 'Match Method', 'Reason Not Identified'):
            w = 34
        ws.column_dimensions[get_column_letter(cidx)].width = w
    ws.sheet_view.showGridLines = False
    return ncols, {src: cidx for cidx, src, _ in fmts}


def fill_cagr(ws, pos, src5, rows):
    """5Y cumulative % -> annualised %, as a live formula."""
    src_col = get_column_letter(pos[src5])
    tgt = pos['__CAGR5__']
    for r in rows:
        ref = f'{src_col}{r}'
        ws.cell(row=r, column=tgt, value=(
            f'=IF(N({ref})=0,IF({ref}="","",0),'
            f'IF({ref}<=-100,"",IFERROR((POWER(1+{ref}/100,1/5)-1)*100,"")))'))


def metric_groups(prefix=''):
    """Financial column groups. prefix='A_' for the merged marquee frame."""
    p = prefix
    return [
        ('Size', '7F6000', [('Market Cap', p + 'Market capitalization', BIG),
                            ('Ccy', p + 'Market capitalization - Currency', TXT)]),
        ('Valuation', GOLD, [('P/E (TTM)', p + 'Price to earnings ratio', X),
                             ('P/E (Forward)', p + 'Price to earnings ratio forward', X)]),
        ('Return Ratios', '1F6B3B', [
            ('ROIC %', p + 'Return on invested capital %, Trailing 12 months', PCT),
            ('ROCE %', p + 'Return on capital employed %, Trailing 12 months', PCT),
            ('ROA %', p + 'Return on assets %, Trailing 12 months', PCT),
            ('ROE %', p + 'Return on equity %, Trailing 12 months', PCT)]),
        ('Profitability Margins', '2F855A', [
            ('Gross Margin %', p + 'Gross margin %, Annual', PCT),
            ('Operating Margin %', p + 'Operating margin %, Annual', PCT),
            ('Net Margin %', p + 'Net margin %, Annual', PCT),
            ('FCF Margin %', p + 'Free cash flow margin %, Annual', PCT)]),
        ('Cash Flow', '116466', [('Free Cash Flow', p + 'Free cash flow, Annual', BIG),
                                 ('FCF Ccy', p + 'Free cash flow, Annual - Currency', TXT)]),
        ('Balance Sheet & Solvency', '7B3F00', [
            ('Current Ratio', p + 'Current ratio, Quarterly', X),
            ('Debt / Equity', p + 'Debt to equity ratio, Quarterly', X),
            ('Interest Coverage', p + 'Interest coverage, Annual', X),
            ('EBITDA Int. Coverage', p + 'EBITDA interest coverage, Annual', X)]),
        ('Quality & Distress Scores', PLUM, [
            ('Altman Z-Score', p + 'Altman Z-score, Trailing 12 months', Z),
            ('Piotroski F-Score', p + 'Piotroski F-score, Trailing 12 months', F)]),
        ('Price Performance', '7A1F3D', [
            ('YTD %', p + 'Performance %, Year to date', PCT),
            ('6M %', p + 'Performance %, 6 months', PCT),
            ('1Y %', p + 'Performance %, 1 year', PCT),
            ('5Y Cumulative %', p + 'Performance %, 5 years', PCT),
            ('5Y Annualised % (CAGR)', '__CAGR5__', PCT)]),
        ('Price & Technicals', '3F3F76', [
            ('Price', p + 'Price', NUM), ('Ccy', p + 'Price - Currency', TXT),
            ('EMA 50', p + 'Exponential moving average, 50, 1 day', NUM),
            ('EMA 200', p + 'Exponential moving average, 200, 1 day', NUM),
            ('RSI 14', p + 'Relative strength index, 14, 1 day', '0.0')]),
    ]


# ------------------------------------------------------------------ sheet builders ---
def sheet_readme(wb, n_matched, n_unid, n_all, screen_counts, stamp):
    ws = wb.active
    ws.title = 'Read Me'
    ws.sheet_view.showGridLines = False
    ws['B2'] = 'Marquee Investor Portfolio and GICS Sector Screener'
    ws['B2'].font = Font(name='Arial', size=16, bold=True, color=NAVY)
    ws['B3'] = (f"Updated as of {stamp['data']}  |  marquee holdings list as of {stamp['marquee']}"
                f"  |  workbook built {stamp['built']} by marquee_screener.py")
    ws['B3'].font = Font(name='Arial', size=10, italic=True, color='595959')

    passes = ', '.join(f'{g} {n}' for g, n in screen_counts.items())
    blocks = [
        ('Contents', [
            ('Marquee Portfolio', f'{n_matched} marquee holdings matched to the universe, financials grouped by discipline.'),
            ('Unidentified in Marquee', f'{n_unid} marquee line items with no company-level financial match, each with a reason.'),
            ('Sector Summary', 'Live formula roll-up of the Marquee Portfolio sheet by TradingView sector.'),
            ('All Stocks', f'{n_all} companies - the full source universe, flagged for marquee ownership.'),
            ('Screen Settings', 'The exact threshold set applied to each GICS sector, with overrides flagged.'),
            ('Screen - All Sectors', 'Every row judged against its own GICS sector block and the results combined, so this sheet is the union of the eleven sector sheets. A GICS column on every row shows which rule set applied. Rows whose TradingView sector maps to no GICS sector are judged on the All Sectors block. Set ALL_SECTORS_USES_SECTOR_RULES to False to apply one market-wide rule instead.'),
            ('Screen - <sector> sheets', 'One sheet per GICS sector, split into four labelled blocks: marquee holdings US, marquee holdings non-US, all other stocks US, all other stocks non-US. Each block is sorted by market cap.'),
        ]),
        ('Column groups', [
            ('Identification', 'Ticker, company, sector, industry, country and how the record was matched.'),
            ('Superinvestor Ownership', 'Aggregate portfolio weight, number of investors holding, largest single weight, average hold price and 52-week context.'),
            ('Size', 'Market capitalisation and its reporting currency.'),
            ('Valuation', 'P/E trailing twelve months and forward P/E.'),
            ('Return Ratios', 'ROIC, ROCE, ROA and ROE - all trailing twelve months.'),
            ('Profitability Margins', 'Gross, operating, net and free cash flow margins - annual.'),
            ('Cash Flow', 'Annual free cash flow in the reporting currency.'),
            ('Balance Sheet & Solvency', 'Current ratio, debt to equity, interest coverage and EBITDA interest coverage.'),
            ('Quality & Distress Scores', 'Altman Z-score for bankruptcy risk - below 1.8 is distress, above 3.0 is safe, and the bands do not apply to banks or insurers. Piotroski F-score out of 9 for fundamental momentum.'),
            ('Price Performance', 'Year to date, six month and one year performance, plus the five year figure both cumulative and annualised.'),
            ('Price & Technicals', 'Last price, 50 and 200 day exponential moving averages and 14 day RSI.'),
        ]),
        ('How the marquee match was made (in priority order)', [
            ('1. Verified cross-listing alias', 'A hand-checked map for ADRs and alternate share classes, for example TSM to 2330 in Taiwan and BUD to ABI in Belgium.'),
            ('2. Exact ticker', 'Symbol present in both files. Where a symbol appears on several exchanges, the row whose price is within 2% of the marquee quote wins, then US registration, then largest market cap.'),
            ('3. Ticker variant', 'Legacy -OLD suffix removed, or dot and underscore share-class separators swapped.'),
            ('4. Company name match', 'Normalised company name equality after stripping legal suffixes and share-class labels.'),
            ('No fuzzy matching', 'Approximate name similarity was tested and rejected - short legal names such as "AB S.A." score as perfect matches against unrelated companies. Anything unconfirmed is reported as unidentified rather than guessed.'),
        ]),
        ('The sector screens', [
            ('GICS mapping', 'TradingView sectors are folded into the eleven GICS sectors. Real Estate has no TradingView sector, so REITs are pulled out of Finance and other sectors by industry name.'),
            ('Unmapped names', 'About 800 companies sit in the TradingView "Miscellaneous" or "Government" sectors, which map to no GICS sector. They appear on All Stocks and can pass the All Sectors screen, but they reach no individual sector sheet. That is why the sector sheets do not sum to the All Sectors count.'),
            ('Two threshold sets', "THRESHOLD_MODE selects one. 'base' gives every sector the identical default thresholds, which is the like-for-like view and what runs unless changed. 'tuned' switches to a set where a few sectors are relaxed because they structurally cannot clear the base rule - banks on ROA, utilities on debt to equity."),
            ('Active set this run', f"THRESHOLD_MODE = '{THRESHOLD_MODE}'. See the Screen Settings sheet for the exact numbers applied to each sector."),
            ('Blank policy', f'BLANK_FAILS is currently {BLANK_FAILS}. ' + ('Missing values are ignored, so a stock is judged only on the metrics it reports.' if not BLANK_FAILS else 'Any stock missing a screened metric is rejected, which is strict given Altman covers about three quarters of the universe.') + f' A stock must also report at least {int(MIN_DATA_COVERAGE*100)}% of the screened metrics to qualify, so thinly covered names cannot pass by default.'),
            ('Passing counts this run', passes),
            ('To change a threshold', "Open marquee_screener.py and edit the sector block inside SECTOR_OVERRIDES (base) or SECTOR_PRESETS (tuned), then run the script again. Sector_Screener.html lets you test values live and copy a ready-made block."),
        ]),
        ('Reading the numbers', [
            ('Ratios are currency neutral', 'P/E, returns, margins and coverage compare directly across rows regardless of listing currency.'),
            ('Absolute figures are not', 'Market capitalisation, free cash flow and price carry their own currency column. Do not sum across currencies without converting.'),
            ('Percentages', 'Stored as the source reported them, so 55.6 means 55.6%.'),
            ('5Y Annualised % (CAGR)', 'A live formula on the cumulative column. Price only, so dividends are excluded, and it assumes a full five years of history. Blank where the source is blank or the loss is 100% or worse.'),
            ('Two as-of dates', 'Ownership columns carry the marquee export date, all financial and price columns the screener export date. Use Price, not Current Price, for anything live.'),
            ('Altman averages are banded', 'Sector Summary averages Altman Z only between -50 and 100 - a few leveraged financials carry scores in the tens of thousands that would swamp the mean. The detail sheets show every value as reported.'),
        ]),
    ]
    r = 5
    for title, items in blocks:
        ws.cell(row=r, column=2, value=title).font = Font(name='Arial', size=12, bold=True, color=GOLD)
        r += 1
        for k, v in items:
            ws.cell(row=r, column=2, value=k).font = Font(name='Arial', size=10, bold=True)
            c = ws.cell(row=r, column=3, value=v)
            c.font = Font(name='Arial', size=10)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[r].height = 28
            r += 1
        r += 1
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 118


def sheet_marquee(wb, matched):
    groups = [('Identification', NAVY, [
        ('Ticker', 'Symbol', TXT), ('Company (Marquee list)', 'Stock', TXT),
        ('Matched Ticker', 'Matched Symbol', TXT), ('Matched Company', 'A_Description', TXT),
        ('Sector', 'A_Sector', TXT), ('GICS Sector', 'GICS Sector', TXT),
        ('Industry', 'A_Industry', TXT),
        ('Country', 'A_Country or region of registration', TXT),
        ('Match Method', 'Match Method', TXT), ('Match Confidence', 'Match Confidence', TXT),
        ('Data Basis', 'Data Basis', TXT)]),
        ('Superinvestor Ownership', '2E5C8A', [
            ('Portfolio Wt %', 'Portfolio %', PCT), ('No. of Investors', 'Ownership Count', BIG),
            ('Max Single Wt %', 'Max %', PCT), ('Avg Hold Price', 'Hold Price', USD),
            ('Current Price (US line)', 'Current Price', USD),
            ('52W Low', '52W Low', USD), ('52W High', '52W High', USD),
            ('% Above 52W Low', '% Above 52W Low', PCT)])] + metric_groups('A_')
    ws = wb.create_sheet('Marquee Portfolio')
    _, pos = write_grid(ws, matched, groups)
    fill_cagr(ws, pos, 'A_Performance %, 5 years', range(3, 3 + len(matched)))
    return ws, pos


def sheet_unidentified(wb, unid):
    groups = [('Identification', NAVY, [('Ticker', 'Symbol', TXT),
                                        ('Company (Marquee list)', 'Stock', TXT)]),
              ('Superinvestor Ownership', '2E5C8A', [
                  ('Portfolio Wt %', 'Portfolio %', PCT),
                  ('No. of Investors', 'Ownership Count', BIG),
                  ('Max Single Wt %', 'Max %', PCT), ('Avg Hold Price', 'Hold Price', USD),
                  ('Current Price', 'Current Price', USD), ('52W Low', '52W Low', USD),
                  ('52W High', '52W High', USD), ('% Above 52W Low', '% Above 52W Low', PCT)]),
              ('Diagnosis', '7A1F3D', [('Reason Not Identified', 'Reason Not Identified', TXT)])]
    ws = wb.create_sheet('Unidentified in Marquee')
    write_grid(ws, unid, groups)


def sheet_sector_summary(wb, matched, pos):
    ws = wb.create_sheet('Sector Summary')
    ws.sheet_view.showGridLines = False
    last = 2 + len(matched)
    sc = get_column_letter(pos['A_Sector'])
    srng = f"'Marquee Portfolio'!${sc}$3:${sc}${last}"

    def rng(src):
        L = get_column_letter(pos[src])
        return f"'Marquee Portfolio'!${L}$3:${L}${last}"

    heads = [('Sector', None), ('Holdings', None), ('Aggregate Portfolio Wt %', 'Portfolio %'),
             ('Avg P/E (TTM)', 'A_Price to earnings ratio'),
             ('Avg ROE %', 'A_Return on equity %, Trailing 12 months'),
             ('Avg ROIC %', 'A_Return on invested capital %, Trailing 12 months'),
             ('Avg Operating Margin %', 'A_Operating margin %, Annual'),
             ('Avg Net Margin %', 'A_Net margin %, Annual'),
             ('Avg 1Y Performance %', 'A_Performance %, 1 year'),
             ('Avg 5Y Annualised %', '__CAGR5__'),
             ('Avg Altman Z (-50 to 100 band)', 'A_Altman Z-score, Trailing 12 months'),
             ('Avg Piotroski F', 'A_Piotroski F-score, Trailing 12 months')]
    ws['A1'] = 'Marquee Portfolio - Sector Roll-up'
    ws['A1'].font = Font(name='Arial', size=13, bold=True, color=NAVY)
    ws['A2'] = ('Every figure is a live formula reading the Marquee Portfolio sheet. Averages are '
                'simple, not weighted, and skip blanks.')
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
    for j, (h, _) in enumerate(heads, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = H2F, PatternFill('solid', fgColor=NAVY), CEN, BORDER

    def fmt_for(h):
        return X if 'P/E' in h else (Z if 'Altman' in h else (F if 'Piotroski' in h else PCT))

    sectors = sorted(s for s in matched['A_Sector'].dropna().unique())
    for i, s in enumerate(sectors):
        r = 5 + i
        ws.cell(row=r, column=1, value=s).font = BF
        ws.cell(row=r, column=2, value=f'=COUNTIFS({srng},$A{r})').number_format = BIG
        for j, (h, src) in enumerate(heads[2:], start=3):
            if j == 3:
                f = f'=IFERROR(SUMIFS({rng(src)},{srng},$A{r}),0)'
            elif 'Altman' in h:
                f = (f'=IFERROR(AVERAGEIFS({rng(src)},{srng},$A{r},{rng(src)},">-50",'
                     f'{rng(src)},"<100"),"")')
            else:
                f = f'=IFERROR(AVERAGEIFS({rng(src)},{srng},$A{r}),"")'
            ws.cell(row=r, column=j, value=f).number_format = fmt_for(h)
        for j in range(1, len(heads) + 1):
            c = ws.cell(row=r, column=j)
            c.font, c.border = BF, BORDER
            if i % 2:
                c.fill = STRIPE
    tr = 5 + len(sectors)
    ws.cell(row=tr, column=1, value='Total / Portfolio').font = BOLD
    ws.cell(row=tr, column=2, value=f'=SUM(B5:B{tr - 1})').number_format = BIG
    ws.cell(row=tr, column=3, value=f'=SUM(C5:C{tr - 1})').number_format = PCT
    for j, (h, src) in enumerate(heads[3:], start=4):
        f = (f'=IFERROR(AVERAGEIFS({rng(src)},{rng(src)},">-50",{rng(src)},"<100"),"")'
             if 'Altman' in h else f'=IFERROR(AVERAGE({rng(src)}),"")')
        ws.cell(row=tr, column=j, value=f).number_format = fmt_for(h)
    for j in range(1, len(heads) + 1):
        c = ws.cell(row=tr, column=j)
        c.font, c.border = BOLD, BORDER
        c.fill = PatternFill('solid', fgColor='FFF2CC')
    ws.column_dimensions['A'].width = 30
    for j in range(2, len(heads) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = 'B5'


def sheet_all_stocks(wb, a, matched):
    keys = set(zip(matched['Matched Symbol'], matched['A_Description']))
    a2 = a.drop(columns=[c for c in a.columns if c.startswith('_')]).copy()
    a2['Marquee Flag'] = ['Marquee holding' if (s, d) in keys else ''
                          for s, d in zip(a['Symbol'], a['Description'])]
    a2['GICS Sector'] = assign_gics(a2)
    a2 = a2.sort_values('Market capitalization', ascending=False).reset_index(drop=True)
    groups = [('Identification', NAVY, [
        ('Ticker', 'Symbol', TXT), ('Company', 'Description', TXT),
        ('Sector', 'Sector', TXT), ('GICS Sector', 'GICS Sector', TXT),
        ('Industry', 'Industry', TXT),
        ('Country', 'Country or region of registration', TXT),
        ('Marquee Flag', 'Marquee Flag', TXT)])] + metric_groups('')
    ws = wb.create_sheet('All Stocks')
    _, pos = write_grid(ws, a2, groups)
    fill_cagr(ws, pos, 'Performance %, 5 years', range(3, 3 + len(a2)))
    return a2


def sheet_screen_settings(wb, screen_counts, stamp):
    ws = wb.create_sheet('Screen Settings')
    ws.sheet_view.showGridLines = False
    ws['A1'] = f"Active screening thresholds by GICS sector   |   updated as of {stamp['data']}"
    ws['A1'].font = Font(name='Arial', size=13, bold=True, color=NAVY)
    src = 'SECTOR_PRESETS (sector-tuned set)' if THRESHOLD_MODE == 'tuned' else \
          'SECTOR_OVERRIDES (base set, every sector identical)'
    ws['A2'] = (f'THRESHOLD_MODE = {THRESHOLD_MODE!r}, so these come from {src}. Edit that block in '
                'marquee_screener.py and run again. "off" means the filter is not applied, and '
                'values in gold differ from DEFAULT_FILTERS.')
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
    keys = list(DEFAULT_FILTERS.keys())
    heads = ['GICS Sector', 'Names passing'] + [FILTER_SPECS[k][2] +
             (' <=' if FILTER_SPECS[k][1] == 'max' else ' >=') for k in keys]
    for j, h in enumerate(heads, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = H2F, PatternFill('solid', fgColor=NAVY), CEN, BORDER
    r = 5
    ws.cell(row=r, column=1, value='DEFAULT (base rules)').font = BOLD
    ws.cell(row=r, column=2, value='-').font = BF
    for j, k in enumerate(keys, start=3):
        v = DEFAULT_FILTERS[k]
        c = ws.cell(row=r, column=j, value='off' if v is None else v)
        c.font = BOLD
    for j in range(1, len(heads) + 1):
        c = ws.cell(row=r, column=j)
        c.border, c.alignment = BORDER, RIGHT if j > 2 else LEFT
        c.fill = PatternFill('solid', fgColor='FFF2CC')
    for gics in screen_order():
        r += 1
        f = sector_filters(gics)
        ov = {k: v for k, v in active_set().get(gics, {}).items() if v != DEFAULT_FILTERS[k]}
        ws.cell(row=r, column=1, value=gics).font = BF
        ws.cell(row=r, column=2, value=screen_counts.get(gics, 0)).number_format = BIG
        for j, k in enumerate(keys, start=3):
            v = f[k]
            c = ws.cell(row=r, column=j, value='off' if v is None else v)
            c.font = Font(name='Arial', size=9, bold=True, color='9C6500') if k in ov else BF
        for j in range(1, len(heads) + 1):
            c = ws.cell(row=r, column=j)
            c.border = BORDER
            c.alignment = RIGHT if j > 2 else LEFT
            if (r - 5) % 2:
                c.fill = STRIPE
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 14
    for j in range(3, len(heads) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.row_dimensions[4].height = 34
    ws.freeze_panes = 'C5'


def sheet_sector_screen(wb, gics, passed, rule_text, universe_label, stamp):
    """One sheet per sector, split into four labelled blocks:
       marquee holdings US, marquee holdings non-US, other stocks US, other stocks non-US.
    """
    ws = wb.create_sheet(SHEET_NAMES[gics])
    ws.sheet_view.showGridLines = False
    title = (f'{gics} - screened names ({universe_label})' if gics != ALL_LABEL else
             f'All sectors combined - screened names ({universe_label})')
    ws['A1'] = f"{title}   |   updated as of {stamp['data']}"
    ws['A1'].font = Font(name='Arial', size=13, bold=True, color=NAVY)
    ws['A2'] = 'Filters: ' + (rule_text or 'no filters applied')
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='595959')
    blk = 'SECTOR_PRESETS' if THRESHOLD_MODE == 'tuned' else 'SECTOR_OVERRIDES'
    ws['A3'] = (f"{len(passed)} names cleared these filters (THRESHOLD_MODE = '{THRESHOLD_MODE}'). "
                f"To change a threshold, edit {blk}['{gics}'] in marquee_screener.py and run again. "
                f"To relax every sector at once, set THRESHOLD_MODE = 'tuned'.")
    ws['A3'].font = Font(name='Arial', size=9, italic=True, color='9C6500')

    groups = [('Identification', NAVY, [
        ('Ticker', 'Symbol', TXT), ('Company', 'Description', TXT),
        ('TradingView Sector', 'Sector', TXT), ('Industry', 'Industry', TXT),
        ('Country', 'Country or region of registration', TXT),
        ('Marquee Flag', 'Marquee Flag', TXT),
        ('Investors Holding', 'Investors Holding', BIG),
        ('Marquee Wt %', 'Marquee Wt %', PCT)])] + metric_groups('')
    if gics == ALL_LABEL:      # the combined sheet needs the sector on every row
        groups[0][2].insert(2, ('GICS Sector', 'GICS Sector', TXT))

    if len(passed) == 0:
        ws['A5'] = ('No stock in this sector cleared these filters. Either set '
                    "THRESHOLD_MODE = 'tuned' for the sector-aware thresholds, or open "
                    f"{blk}['{gics}'] and loosen the P/E, ROA or Altman value. "
                    'Sector_Screener.html lets you test values live before committing them.')
        ws['A5'].font = Font(name='Arial', size=11, bold=True, color='C00000')
        ws.column_dimensions['A'].width = 120
        return

    # header only, then the four blocks are written by hand underneath
    ncols, pos = write_grid(ws, passed.iloc[0:0], groups, freeze_col=3, start_row=5)
    flat = [(pos[src], src, fmt) for _, _, cols in groups for (_, src, fmt) in cols]

    is_mq = passed['Marquee Flag'] == 'Marquee holding'
    is_us = passed['Country or region of registration'] == 'United States'
    sections = [
        ('MARQUEE HOLDINGS  -  US LISTED',            passed[is_mq & is_us],   '2E5C8A'),
        ('MARQUEE HOLDINGS  -  NON-US LISTED',        passed[is_mq & ~is_us],  '3F6FA5'),
        ('ALL OTHER STOCKS  -  US LISTED',            passed[~is_mq & is_us],  '7B3F00'),
        ('ALL OTHER STOCKS  -  NON-US LISTED',        passed[~is_mq & ~is_us], '9E6B33'),
    ]

    r = 7
    data_rows = []
    for title, block, colour in sections:
        block = block.sort_values('Market capitalization', ascending=False)
        for j in range(1, ncols + 1):
            c = ws.cell(row=r, column=j)
            c.fill = PatternFill('solid', fgColor=colour)
            c.border = BORDER
        head = ws.cell(row=r, column=1, value=f'{title}   ({len(block)} names)')
        head.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        head.alignment = LEFT
        r += 1
        if len(block) == 0:
            c = ws.cell(row=r, column=1, value='none cleared the filters in this group')
            c.font = Font(name='Arial', size=9, italic=True, color='808080')
            for j in range(1, ncols + 1):
                ws.cell(row=r, column=j).border = BORDER
            r += 1
            continue
        for i, (_, row) in enumerate(block.iterrows()):
            for cidx, src, fmt in flat:
                c = ws.cell(row=r, column=cidx)
                if not src.startswith('__'):
                    v = row.get(src, None)
                    if isinstance(v, float) and pd.isna(v):
                        v = None
                    if isinstance(v, str) and v.strip() == '':
                        v = None
                    c.value = v
                c.font, c.number_format, c.border = BF, fmt, BORDER
                c.alignment = LEFT if fmt == TXT else RIGHT
                if i % 2:
                    c.fill = STRIPE
            data_rows.append(r)
            r += 1

    ws.auto_filter.ref = f"A6:{get_column_letter(ncols)}{r - 1}"
    fill_cagr(ws, pos, 'Performance %, 5 years', data_rows)




# ------------------------------------------------------------------ deep app feed -----
# The deep analysis app needs the TradingView numbers, and it cannot get them from the
# HTML screener. So the build publishes two small JSON files next to index.html:
#
#   stocks.json        keyed by ISIN, one entry per name worth analysing. Carries every
#                      TradingView field plus the Dataroma ownership, so the deep page
#                      shows TradingView as its primary source and reaches for Yahoo only
#                      where TradingView has nothing.
#   sector_stats.json  per GICS sector quartiles for the main ratios. This is what turns
#                      "ROE 18.2%" into "ROE 18.2%, sector median 12.8%, 72nd percentile",
#                      and the data was already sitting in the export.
#
# Only screened names and marquee holdings go in, not all 30,000 rows. That keeps the file
# near a megabyte instead of ten, which matters because the app fetches it on load.

DEEP_FEED_FIELDS = [
    ('Symbol', 'ticker'), ('Description', 'name'), ('ISIN', 'isin'),
    ('Sector', 'tv_sector'), ('GICS Sector', 'gics'), ('Industry', 'industry'),
    ('Country or region of registration', 'country'),
    ('Market capitalization', 'mcap'), ('Market capitalization - Currency', 'mcap_ccy'),
    ('Price', 'price'), ('Price - Currency', 'ccy'),
    ('Price to earnings ratio', 'pe'), ('Price to earnings ratio forward', 'fpe'),
    ('Return on invested capital %, Trailing 12 months', 'roic'),
    ('Return on capital employed %, Trailing 12 months', 'roce'),
    ('Return on assets %, Trailing 12 months', 'roa'),
    ('Return on equity %, Trailing 12 months', 'roe'),
    ('Gross margin %, Annual', 'gross_margin'),
    ('Operating margin %, Annual', 'op_margin'),
    ('Net margin %, Annual', 'net_margin'),
    ('Free cash flow margin %, Annual', 'fcf_margin'),
    ('Free cash flow, Annual', 'fcf'), ('Free cash flow, Annual - Currency', 'fcf_ccy'),
    ('Current ratio, Quarterly', 'current_ratio'),
    ('Debt to equity ratio, Quarterly', 'debt_equity'),
    ('Interest coverage, Annual', 'interest_cover'),
    ('EBITDA interest coverage, Annual', 'ebitda_cover'),
    ('Altman Z-score, Trailing 12 months', 'altman_z'),
    ('Piotroski F-score, Trailing 12 months', 'piotroski_f'),
    ('Performance %, Year to date', 'perf_ytd'),
    ('Performance %, 6 months', 'perf_6m'),
    ('Performance %, 1 year', 'perf_1y'),
    ('Performance %, 5 years', 'perf_5y'),
    ('Exponential moving average, 50, 1 day', 'ema50'),
    ('Exponential moving average, 200, 1 day', 'ema200'),
    ('Relative strength index, 14, 1 day', 'rsi14'),
]

# the period each TradingView field actually covers, so the deep page can label it and
# nobody compares a trailing-twelve-month return against an annual margin by accident
DEEP_FEED_PERIODS = {
    'pe': 'TTM', 'fpe': 'Forward', 'roic': 'TTM', 'roce': 'TTM', 'roa': 'TTM', 'roe': 'TTM',
    'gross_margin': 'Annual', 'op_margin': 'Annual', 'net_margin': 'Annual',
    'fcf_margin': 'Annual', 'fcf': 'Annual',
    'current_ratio': 'Quarterly', 'debt_equity': 'Quarterly',
    'interest_cover': 'Annual', 'ebitda_cover': 'Annual',
    'altman_z': 'TTM', 'piotroski_f': 'TTM',
    'mcap': 'Current', 'price': 'Current', 'ema50': 'Current', 'ema200': 'Current',
    'rsi14': 'Current',
}

# ratios worth a peer distribution; absolute figures are excluded because a median market
# cap or free cash flow across a sector says nothing useful
PEER_METRICS = ['pe', 'fpe', 'roic', 'roce', 'roa', 'roe', 'gross_margin', 'op_margin',
                'net_margin', 'fcf_margin', 'current_ratio', 'debt_equity',
                'interest_cover', 'altman_z', 'piotroski_f']


def _clean_number(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            value = float(value)
        except ValueError:
            return value
    if isinstance(value, float):
        if pd.isna(value):
            return None
        return round(value, 4)
    return value


def write_deep_feed(all_rows, screened_symbols, out_dir, publish_dir, stamp):
    """Write stocks.json and sector_stats.json for the deep analysis app."""
    rows = all_rows[
        all_rows['Symbol'].isin(screened_symbols)
        | (all_rows['Marquee Flag'] == 'Marquee holding')
    ].copy()
    rows = rows[rows['ISIN'].notna()]
    if len(rows) == 0:
        print('      deep feed skipped - no screened rows carry an ISIN')
        return

    stocks = {}
    for _, r in rows.iterrows():
        entry = {}
        for src, key in DEEP_FEED_FIELDS:
            entry[key] = _clean_number(r.get(src))
        entry['marquee'] = bool(r.get('Marquee Flag') == 'Marquee holding')
        entry['marquee_weight_pct'] = _clean_number(r.get('Marquee Wt %'))
        entry['marquee_investors'] = _clean_number(r.get('Investors Holding'))
        entry['screened'] = bool(r['Symbol'] in screened_symbols)
        # last writer wins on a duplicate ISIN, which only happens for dual listings
        stocks[r['ISIN']] = entry

    feed = {'as_of': stamp['data'], 'marquee_as_of': stamp['marquee'],
            'built': stamp['built'], 'source': 'TradingView screener export',
            'periods': DEEP_FEED_PERIODS, 'count': len(stocks), 'stocks': stocks}

    # peer distributions, computed on the FULL universe rather than the screened subset,
    # because a median taken from names that already passed a quality screen is not a
    # sector median - it is a median of the winners
    stats = {}
    for gics in SECTOR_MAP:
        sub = all_rows[all_rows['GICS Sector'] == gics]
        if len(sub) < 20:
            continue
        block = {'count': int(len(sub))}
        for src, key in DEEP_FEED_FIELDS:
            if key not in PEER_METRICS:
                continue
            series = pd.to_numeric(sub[src], errors='coerce').dropna()
            if len(series) < 20:
                continue
            block[key] = {'n': int(len(series)),
                          'p25': round(float(series.quantile(0.25)), 4),
                          'median': round(float(series.median()), 4),
                          'p75': round(float(series.quantile(0.75)), 4)}
        stats[gics] = block

    stats_feed = {'as_of': stamp['data'], 'basis': 'full TradingView universe, not the '
                                                   'screened subset', 'sectors': stats}

    for folder in [d for d in (out_dir, publish_dir) if d]:
        os.makedirs(folder, exist_ok=True)
        with open(os.path.join(folder, 'stocks.json'), 'w') as fh:
            json.dump(feed, fh, separators=(',', ':'))
        with open(os.path.join(folder, 'sector_stats.json'), 'w') as fh:
            json.dump(stats_feed, fh, indent=1)
    size = os.path.getsize(os.path.join(out_dir, 'stocks.json')) / 1024
    print(f'      deep feed -> stocks.json ({len(stocks)} names, {size:,.0f} KB) '
          f'and sector_stats.json ({len(stats)} sectors)')


# ------------------------------------------------------------------ html -------------
HTML_COLS = [('Symbol', 'sym'), ('Description', 'name'), ('GICS Sector', 'gics'),
             ('ISIN', 'isin'),
             ('Country or region of registration', 'ctry'), ('Marquee Flag', 'mq'),
             ('Marquee Wt %', 'wt'), ('Market capitalization', 'mcap'),
             ('Price to earnings ratio', 'pe'), ('Price to earnings ratio forward', 'fpe'),
             ('Return on invested capital %, Trailing 12 months', 'roic'),
             ('Return on capital employed %, Trailing 12 months', 'roce'),
             ('Return on assets %, Trailing 12 months', 'roa'),
             ('Return on equity %, Trailing 12 months', 'roe'),
             ('Gross margin %, Annual', 'gm'), ('Operating margin %, Annual', 'om'),
             ('Net margin %, Annual', 'nm'), ('Free cash flow margin %, Annual', 'fcfm'),
             ('Current ratio, Quarterly', 'cr'), ('Debt to equity ratio, Quarterly', 'de'),
             ('Interest coverage, Annual', 'ic'),
             ('Altman Z-score, Trailing 12 months', 'az'),
             ('Piotroski F-score, Trailing 12 months', 'pf'),
             ('Performance %, Year to date', 'ytd'),
             ('Performance %, 1 year', 'y1'), ('Performance %, 5 years', 'y5'),
             ('Price', 'px'), ('Price - Currency', 'ccy')]

# HTML filter key -> (data field, direction, label)
HTML_FILTERS = [(k, HTML_KEY, FILTER_SPECS[k][1], FILTER_SPECS[k][2]) for k, HTML_KEY in [
    ('pe_max', 'pe'), ('roe_min', 'roe'), ('roic_min', 'roic'), ('roce_min', 'roce'),
    ('roa_min', 'roa'), ('gross_margin_min', 'gm'), ('op_margin_min', 'om'),
    ('net_margin_min', 'nm'), ('fcf_margin_min', 'fcfm'), ('altman_min', 'az'),
    ('piotroski_min', 'pf'), ('current_ratio_min', 'cr'), ('de_max', 'de'),
    ('int_cov_min', 'ic'), ('mcap_min', 'mcap')]]


def build_html(a2, out_path, stamp, split=False):
    """Write the screener. split=True writes a light index.html plus a separate data.json,
    which is what a website wants: the page paints immediately and the data streams in."""
    rows = []
    for _, r in a2.iterrows():
        rec = []
        for src, _ in HTML_COLS:
            v = r.get(src, None)
            if src == 'GICS Sector':
                v = None if pd.isna(v) else list(SECTOR_MAP).index(v)
            elif src == 'Marquee Flag':
                v = 1 if v == 'Marquee holding' else 0
            elif isinstance(v, float):
                # four decimals, not two: rounding at the second decimal shifts borderline
                # rows across a threshold and makes the HTML count drift from the workbook
                v = None if pd.isna(v) else round(v, 4)
            elif isinstance(v, str):
                v = v.strip() or None
            elif pd.isna(v):
                v = None
            rec.append(v)
        rows.append(rec)
    payload = {
        'fields': [k for _, k in HTML_COLS],
        'rows': rows,
        'sectors': list(SECTOR_MAP.keys()),
        'dropdown': [ALL_LABEL] + list(SECTOR_MAP.keys()),
        'allLabel': ALL_LABEL,
        'defaults': DEFAULT_FILTERS,
        'base': SECTOR_OVERRIDES,
        'tuned': SECTOR_PRESETS,
        'mode': THRESHOLD_MODE,
        'perSector': ALL_SECTORS_USES_SECTOR_RULES,
        'filters': [{'key': k, 'field': f, 'dir': d, 'label': l} for k, f, d, l in HTML_FILTERS],
        'blankFails': BLANK_FAILS,
        'minCoverage': MIN_DATA_COVERAGE,
        'stamp': stamp,
        'deepUrl': DEEP_APP_URL.rstrip('/'),
        'xlsxUrl': (f'https://github.com/{GITHUB_REPO}/raw/{GITHUB_BRANCH}/{GITHUB_XLSX}'
                    if (GITHUB_UPLOAD and GITHUB_XLSX) else ''),
    }
    blob = json.dumps(payload, separators=(',', ':'))
    if split:
        data_path = os.path.join(os.path.dirname(out_path) or '.', 'data.json')
        with open(data_path, 'w') as fh:
            fh.write(blob)
        bootstrap = ("""document.getElementById('msg').innerHTML =
  '<div class="note" style="padding:10px 12px">Loading the stock data, one moment...</div>';
fetch('data.json' + (location.search||''), {cache:'no-cache'})
  .then(r=>{ if(!r.ok) throw new Error('HTTP '+r.status); return r.json(); })
  .then(d=>{ document.getElementById('msg').innerHTML=''; boot(d); })
  .catch(e=>{ document.getElementById('msg').innerHTML =
    '<div class="warn">Could not load data.json (' + e.message + '). It must sit in the same '
    + 'folder as this page. Opening this file straight from your desktop will not work - '
    + 'browsers block local file reads. Use the single-file version for that.</div>'; });""")
        html = HTML_TEMPLATE.replace('/*__BOOT__*/', bootstrap)
        with open(out_path, 'w') as fh:
            fh.write(html)
        return os.path.getsize(out_path), os.path.getsize(data_path)
    html = HTML_TEMPLATE.replace('/*__BOOT__*/', 'boot(' + blob + ');')
    with open(out_path, 'w') as fh:
        fh.write(html)
    return os.path.getsize(out_path), 0


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>GICS Sector Screener</title>
<style>
 :root{--navy:#1f3864;--gold:#bf8f00;--plum:#4b2e83;--line:#d8dee9;--bg:#f6f7fa;}
 *{box-sizing:border-box}
 body{margin:0;font:14px/1.45 -apple-system,Segoe UI,Roboto,Arial,sans-serif;background:var(--bg);color:#1a1a1a}
 header{background:var(--navy);color:#fff;padding:14px 20px}
 header h1{margin:0;font-size:18px;font-weight:600}
 header p{margin:4px 0 0;font-size:12px;opacity:.8}
 .wrap{display:flex;align-items:flex-start;gap:16px;padding:16px;flex-wrap:wrap}
 .panel{background:#fff;border:1px solid var(--line);border-radius:8px;padding:14px;width:320px;flex:0 0 320px}
 .main{background:#fff;border:1px solid var(--line);border-radius:8px;padding:14px;flex:1 1 640px;min-width:0}
 h2{font-size:13px;text-transform:uppercase;letter-spacing:.05em;color:var(--navy);margin:0 0 10px}
 label{display:block;font-size:12px;color:#444;margin:10px 0 3px}
 select,input[type=number],input[type=text]{width:100%;padding:6px 8px;border:1px solid var(--line);border-radius:5px;font:13px inherit;background:#fff}
 .frow{display:flex;align-items:center;gap:6px;padding:5px 0;border-bottom:1px dashed #eef0f4}
 .frow .nm{flex:1 1 auto;font-size:12px}
 .frow .op{color:var(--gold);font-weight:700;font-size:12px;width:24px;text-align:center}
 .frow input[type=number]{width:78px;padding:4px 6px}
 .frow input[type=checkbox]{accent-color:var(--navy)}
 .btns{display:flex;gap:8px;margin-top:12px;flex-wrap:wrap}
 button{padding:7px 11px;border:1px solid var(--navy);background:var(--navy);color:#fff;border-radius:5px;font:12px inherit;cursor:pointer}
 button.ghost{background:#fff;color:var(--navy)}
 button.gold{background:var(--gold);border-color:var(--gold)}
 .count{font-size:12px;color:#555;margin:8px 0 10px}
 .count b{color:var(--navy);font-size:18px}
 .pill{display:inline-block;background:#eef2f8;border:1px solid var(--line);border-radius:10px;padding:1px 7px;font-size:11px;margin-right:4px}
 .tw{overflow:auto;max-height:66vh;border:1px solid var(--line);border-radius:6px}
 table{border-collapse:collapse;width:100%;font-size:12px}
 th,td{padding:5px 7px;border-bottom:1px solid #eef0f4;white-space:nowrap;text-align:right}
 th{position:sticky;top:0;background:var(--navy);color:#fff;cursor:pointer;font-weight:600;z-index:2}
 th:first-child,td:first-child,th:nth-child(2),td:nth-child(2){text-align:left}
 tbody tr:nth-child(even){background:#f8fafc}
 tbody tr:hover{background:#fff7e0}
 .mqdot{color:var(--gold);font-weight:700}
 td a.tk{color:var(--navy);font-weight:600;text-decoration:none;border-bottom:1px dotted var(--navy)}
 td a.tk:hover{color:var(--gold);border-bottom-color:var(--gold)}
 .warn{background:#fff4f4;border:1px solid #f0c8c8;color:#9c2b2b;padding:9px 11px;border-radius:6px;font-size:12px;margin-bottom:10px}
 .note{font-size:11px;color:#777;margin-top:8px;line-height:1.5}
 .grp{margin-top:14px;padding-top:8px;border-top:2px solid var(--navy)}
</style></head><body>
<header>
 <h1>GICS Sector Screener <span id="asof" style="font-weight:400;opacity:.85;font-size:13px"></span></h1>
 <p>Threshold set switches between the identical base rules and the sector-tuned ones. Change anything here. Once you settle on a threshold set, use Copy as Python and paste it into SECTOR_OVERRIDES in marquee_screener.py.</p>
</header>
<div class="wrap">
 <div class="panel">
  <h2>Sector &amp; region</h2>
  <label>GICS sector</label>
  <select id="sector"></select>
  <label>Region</label>
  <select id="region"><option value="all">US + non-US</option><option value="us">US only</option><option value="non">Non-US only</option></select>
  <label>Universe</label>
  <select id="mqonly"><option value="0">All stocks</option><option value="1">Marquee holdings only</option></select>
  <label>Threshold set</label>
  <select id="preset">
    <option value="base">Base - every sector identical</option>
    <option value="tuned">Sector-tuned - relaxed where needed</option>
  </select>
  <label style="margin-top:8px"><input type="checkbox" id="persector"> Judge each row on its own sector's rules</label>
  <label>Missing values</label>
  <select id="blank"><option value="0">Ignore blanks (loose)</option><option value="1">Blank = reject (strict)</option></select>
  <label>Search name or ticker</label>
  <input type="text" id="q" placeholder="e.g. bank, TSM">

  <div class="grp"><h2>Thresholds</h2>
   <div id="filters"></div>
   <div class="btns">
    <button class="ghost" id="reset">Reset to set</button>
    <button class="ghost" id="loosen">Loosen 25%</button>
    <button class="ghost" id="allof">All off</button>
    <button class="gold" id="copy">Copy as Python</button>
   </div>
   <div class="note">An unchecked box means that filter is not applied. Copy as Python puts a ready-made SECTOR_OVERRIDES block on the clipboard, ready to paste straight into the script.</div>
  </div>
 </div>

 <div class="main">
  <div id="msg"></div>
  <div class="count"><b id="n">0</b> names clear these filters <span id="split"></span></div>
  <div class="btns" style="margin-top:0">
   <button id="csv">Download CSV</button>
   <a id="xlsx" href="#" style="display:none"><button class="ghost">Download full workbook</button></a>
   <span class="pill" id="showing"></span>
  </div>
  <div class="tw"><table><thead><tr id="thead"></tr></thead><tbody id="tbody"></tbody></table></div>
  <div class="note" id="hint">Click a column heading to sort. The table shows the first 400 rows; the CSV download contains the full list.</div>
 </div>
</div>
<script>
let DATA, F, ROWS, SEC; const IX = {};
function boot(d){
  DATA=d; F=DATA.fields; F.forEach((f,i)=>IX[f]=i); ROWS=DATA.rows; SEC=DATA.sectors; init();
}
const COLS = [['sym','Ticker'],['name','Company'],['gicsName','GICS'],['ctry','Country'],['mq','Marquee'],['wt','Wt %'],
 ['mcap','Mkt cap'],['pe','P/E'],['fpe','Fwd P/E'],['roic','ROIC'],['roce','ROCE'],['roa','ROA'],['roe','ROE'],
 ['gm','Gross %'],['om','Op %'],['nm','Net %'],['fcfm','FCF %'],['cr','Cur ratio'],['de','D/E'],
 ['ic','Int cov'],['az','Altman Z'],['pf','Piotroski'],['ytd','YTD %'],['y1','1Y %'],['y5','5Y cum %'],['cagr','5Y ann %'],['px','Price'],['ccy','Ccy']];
let sortKey='mcap', sortDir=-1, cur={};

function sectorDefaults(g){
  const set = document.getElementById('preset').value==='tuned'? DATA.tuned : DATA.base;
  return Object.assign({}, DATA.defaults, set[g]||{});
}
function buildSectorSelect(){
  const s=document.getElementById('sector');
  DATA.dropdown.forEach(g=>{const o=document.createElement('option');o.value=g;o.textContent=g;s.appendChild(o)});
  s.value=DATA.dropdown[0];
}
function buildFilters(){
  const box=document.getElementById('filters'); box.innerHTML='';
  const d=sectorDefaults(document.getElementById('sector').value);
  DATA.filters.forEach(f=>{
    const v=d[f.key], on=(v!==null&&v!==undefined);
    const row=document.createElement('div'); row.className='frow';
    row.innerHTML='<input type="checkbox" '+(on?'checked':'')+' data-k="'+f.key+'">'+
      '<span class="nm">'+f.label+'</span><span class="op">'+(f.dir==='min'?'&ge;':'&le;')+'</span>'+
      '<input type="number" step="any" data-v="'+f.key+'" value="'+(on?v:'')+'">';
    box.appendChild(row);
  });
  box.querySelectorAll('input').forEach(el=>el.addEventListener('input',run));
}
function currentSet(){
  return document.getElementById('preset').value==='tuned'? DATA.tuned : DATA.base;
}
function blockFilters(g){
  // the threshold list for one sector, taken from the selected set rather than the UI boxes
  const d=Object.assign({}, DATA.defaults, currentSet()[g]||{});
  const out=[];
  DATA.filters.forEach(f=>{ const v=d[f.key];
    if(v!==null&&v!==undefined) out.push({field:f.field,dir:f.dir,val:v}); });
  return out;
}
function activeFilters(){
  const out=[];
  document.querySelectorAll('#filters .frow').forEach(r=>{
    const cb=r.querySelector('input[type=checkbox]'), nb=r.querySelector('input[type=number]');
    if(!cb.checked||nb.value==='') return;
    const spec=DATA.filters.find(f=>f.key===cb.dataset.k);
    out.push({field:spec.field,dir:spec.dir,val:parseFloat(nb.value),label:spec.label,key:spec.key});
  });
  return out;
}
function cagr(y5){ if(y5===null||y5===undefined) return null; if(y5<=-100) return null;
  return (Math.pow(1+y5/100,0.2)-1)*100; }

function run(){
  const g=document.getElementById('sector').value, reg=document.getElementById('region').value;
  const mq=document.getElementById('mqonly').value==='1';
  const strict=document.getElementById('blank').value==='1';
  const q=document.getElementById('q').value.trim().toLowerCase();
  const perSector = document.getElementById('persector').checked && g===DATA.allLabel;
  const fs=activeFilters();
  const cache={};
  const rulesFor=sec=>(cache[sec]=cache[sec]||blockFilters(sec in currentSet()? sec : DATA.allLabel));
  let out=[];
  for(const r of ROWS){
    if(g!==DATA.allLabel && SEC[r[IX.gics]]!==g) continue;
    const isUS = r[IX.ctry]==='United States';
    if(reg==='us'&&!isUS) continue;
    if(reg==='non'&&isUS) continue;
    if(mq && !r[IX.mq]) continue;
    if(q && !((r[IX.sym]||'').toLowerCase().includes(q)||(r[IX.name]||'').toLowerCase().includes(q))) continue;
    let ok=true, reported=0;
    const rules = perSector? rulesFor(SEC[r[IX.gics]]||DATA.allLabel) : fs;
    for(const f of rules){
      const v=r[IX[f.field]];
      if(v===null||v===undefined){ if(strict){ok=false;break;} else continue; }
      reported++;
      if(f.dir==='min'? v< f.val : v> f.val){ok=false;break;}
    }
    // same data-coverage guard the script applies, so the counts match the workbook
    if(ok && !strict && rules.length && DATA.minCoverage>0
       && reported/rules.length < DATA.minCoverage) ok=false;
    if(ok) out.push(r);
  }
  cur=out;
  const nus=out.filter(r=>r[IX.ctry]==='United States').length;
  document.getElementById('n').textContent=out.length;
  document.getElementById('split').innerHTML=out.length?
    '<span class="pill">US '+nus+'</span><span class="pill">Non-US '+(out.length-nus)+'</span>':'';
  const ps=document.getElementById('persector').checked;
  document.getElementById('msg').innerHTML =
    (ps && g===DATA.allLabel ? '<div class="note" style="background:#eef2f8;border:1px solid #d8dee9;padding:8px 10px;border-radius:6px;margin-bottom:8px">Each row is being judged on its own sector\'s block from the selected set, so the threshold boxes below are ignored. This matches the Screen - All Sectors sheet in the workbook.</div>' : '')
    + (ps && g!==DATA.allLabel ? '<div class="note" style="background:#fff7e0;border:1px solid #f0dca8;padding:8px 10px;border-radius:6px;margin-bottom:8px">Per-sector rules only apply when the sector is set to All Sectors.</div>' : '')
    + (out.length? '' :
    '<div class="warn">Nothing cleared these filters. Loosen a threshold - try Loosen 25%, or uncheck ROA and Altman. Banks and utilities never clear an ROA of 10% or debt to equity below 2.</div>');
  render();
}
function val(r,k){ if(k==='cagr') return cagr(r[IX.y5]);
  if(k==='gicsName') return SEC[r[IX.gics]]||'-';
  return r[IX[k]]; }
function render(){
  const th=document.getElementById('thead'); th.innerHTML='';
  COLS.forEach(([k,l])=>{const e=document.createElement('th');e.textContent=l+(sortKey===k?(sortDir<0?' \u25bc':' \u25b2'):'');
    e.onclick=()=>{ if(sortKey===k) sortDir=-sortDir; else {sortKey=k;sortDir=-1;} render(); }; th.appendChild(e);});
  const rows=[...cur].sort((x,y)=>{
    let a=val(x,sortKey), b=val(y,sortKey);
    if(a===null||a===undefined) return 1; if(b===null||b===undefined) return -1;
    if(typeof a==='string') return a.localeCompare(b)*(sortDir<0?-1:1);
    return (a-b)*(sortDir<0?-1:1);
  });
  const shown=rows.slice(0,400);
  document.getElementById('showing').textContent='showing '+shown.length+' of '+rows.length;
  const tb=document.getElementById('tbody'); tb.innerHTML='';
  for(const r of shown){
    const tr=document.createElement('tr');
    tr.innerHTML=COLS.map(([k])=>{
      let v=val(r,k);
      if(k==='sym' && DATA.deepUrl){
        // The ISIN goes first because Yahoo's own search resolves it to the right symbol,
        // which beats guessing a venue suffix. The currency comes along so the app can
        // prefer the local listing over a US ADR. The ticker rides as a fallback.
        let q='?ticker='+encodeURIComponent(v);
        const isin=r[IX.isin], ccy=r[IX.ccy];
        if(isin) q+='&isin='+encodeURIComponent(isin);
        if(ccy)  q+='&ccy='+encodeURIComponent(ccy);
        return '<td><a class="tk" target="_blank" rel="noopener" href="'+DATA.deepUrl+'/'+q+
               '" title="Open the deep analysis for '+v+(isin?' ('+isin+')':'')+'">'+v+'</a></td>';
      }
      if(k==='mq') return '<td class="mqdot">'+(v?'\u2605':'')+'</td>';
      if(v===null||v===undefined) return '<td>-</td>';
      if(typeof v==='number'){
        if(k==='mcap') return '<td>'+Math.round(v).toLocaleString()+'</td>';
        return '<td>'+v.toFixed(2)+'</td>';
      }
      return '<td>'+v+'</td>';
    }).join('');
    tb.appendChild(tr);
  }
}
document.getElementById('csv').onclick=()=>{
  const head=COLS.map(c=>c[1]).join(',');
  const body=cur.map(r=>COLS.map(([k])=>{
    let v=val(r,k); if(v===null||v===undefined) return '';
    if(typeof v==='string'&&/[",]/.test(v)) return '"'+v.replace(/"/g,'""')+'"';
    return v;
  }).join(',')).join('\n');
  const b=new Blob([head+'\n'+body],{type:'text/csv'});
  const a=document.createElement('a');a.href=URL.createObjectURL(b);
  a.download='screen_'+document.getElementById('sector').value.replace(/\s+/g,'_')+'_'+DATA.stamp.iso+'.csv';a.click();
};
document.getElementById('copy').onclick=()=>{
  const g=document.getElementById('sector').value;
  const lines=activeFilters().map(f=>"        '"+f.key+"'"+' '.repeat(Math.max(1,20-f.key.length))+': '+f.val+',');
  const offs=DATA.filters.filter(f=>!activeFilters().some(a=>a.key===f.key))
    .map(f=>"        '"+f.key+"'"+' '.repeat(Math.max(1,20-f.key.length))+": None,");
  const txt="    '"+g+"': {\n"+lines.concat(offs).join('\n')+"\n    },";
  navigator.clipboard.writeText(txt).then(
    ()=>alert('Copied. Paste this inside '+(document.getElementById('preset').value==='tuned'?'SECTOR_PRESETS':'SECTOR_OVERRIDES')+' in marquee_screener.py:\n\n'+txt),
    ()=>alert('Clipboard blocked, copy this manually:\n\n'+txt));
};
document.getElementById('reset').onclick=()=>{buildFilters();run();};
document.getElementById('allof').onclick=()=>{
  document.querySelectorAll('#filters input[type=checkbox]').forEach(c=>c.checked=false); run();};
document.getElementById('loosen').onclick=()=>{
  document.querySelectorAll('#filters .frow').forEach(r=>{
    const cb=r.querySelector('input[type=checkbox]'), nb=r.querySelector('input[type=number]');
    if(!cb.checked||nb.value==='') return;
    const spec=DATA.filters.find(f=>f.key===cb.dataset.k);
    const v=parseFloat(nb.value);
    nb.value=+(spec.dir==='min'? v*0.75 : v*1.25).toFixed(2);
  }); run();};
['sector','region','mqonly','blank','preset','persector'].forEach(id=>document.getElementById(id).addEventListener('change',()=>{
  if(id==='sector'||id==='preset') buildFilters(); run();}));
document.getElementById('q').addEventListener('input',run);
function init(){
  document.getElementById('asof').textContent =
    '  -  updated as of ' + DATA.stamp.data + '  (marquee list ' + DATA.stamp.marquee + ')';
  if(DATA.deepUrl){
    document.getElementById('hint').innerHTML =
      'Click a <b>ticker</b> to open the live deep analysis for that company. Click a column heading to sort. '
      + 'The table shows the first 400 rows; the CSV download contains the full list.';
  }
  if(DATA.xlsxUrl){ const a=document.getElementById('xlsx');
    a.href=DATA.xlsxUrl; a.style.display='inline-block'; }
  buildSectorSelect();
  document.getElementById('preset').value = DATA.mode;
  document.getElementById('persector').checked = DATA.perSector;
  document.getElementById('blank').value = DATA.blankFails? '1':'0';
  buildFilters();
  run();
}
/*__BOOT__*/
</script></body></html>
"""




def _gh(method, url, token, payload=None):
    """One GitHub API call. Returns the parsed JSON body."""
    import urllib.error, urllib.request
    data = json.dumps(payload).encode() if payload is not None else None
    req = urllib.request.Request(url, data=data, method=method, headers={
        'Authorization': f'Bearer {token}',
        'Accept': 'application/vnd.github+json',
        'X-GitHub-Api-Version': '2022-11-28',
        'Content-Type': 'application/json',
        'User-Agent': 'marquee-screener',
    })
    try:
        with urllib.request.urlopen(req) as r:
            return json.loads(r.read() or b'{}')
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors='replace')[:400]
        hint = {
            401: 'the token is wrong, expired, or was never set',
            403: 'the token lacks Contents: Read and write on this repository',
            404: 'repo, branch or path not found, or the token cannot see this repo',
            409: 'the branch moved while uploading - just run again',
            422: 'GitHub rejected the content, often a stale sha - just run again',
        }.get(e.code, 'see the message above')
        sys.exit(f'\n[ERROR] GitHub said {e.code} on {method} {url}\n'
                 f'        Likely cause: {hint}\n        Response: {body}\n')
    except urllib.error.URLError as e:
        sys.exit(f'\n[ERROR] Could not reach GitHub: {e.reason}\n'
                 f'        Check the internet connection or a corporate proxy.\n')


def publish_to_github(files, stamp):
    """Commit several files to a GitHub repo in one commit, using the Git Data API.

    files is a list of (local path, path inside the repo).
    blob -> tree -> commit -> ref is how git itself works, so a 6 MB file is no problem.
    The simpler Contents API is unreliable above about 1 MB.
    """
    import base64
    token = (GITHUB_TOKEN or os.environ.get('GITHUB_TOKEN', '')).strip()
    if not token or 'PASTE_YOUR' in token:
        print('      [WARN] no token found, upload skipped. Fill GITHUB_TOKEN or set the '
              'GITHUB_TOKEN environment variable.')
        return
    files = [(lp, rp) for lp, rp in files if rp and os.path.exists(lp)]
    if not files:
        print('      [WARN] nothing to upload')
        return

    api = f'https://api.github.com/repos/{GITHUB_REPO}'
    total = sum(os.path.getsize(lp) for lp, _ in files) / 1024 / 1024
    names = ', '.join(rp for _, rp in files)
    print(f'      uploading {total:,.1f} MB to {GITHUB_REPO} ({GITHUB_BRANCH}): {names}')

    ref = _gh('GET', f'{api}/git/ref/heads/{GITHUB_BRANCH}', token)
    parent = ref['object']['sha']
    base_tree = _gh('GET', f'{api}/git/commits/{parent}', token)['tree']['sha']

    tree_items = []
    for local, repo_path in files:
        with open(local, 'rb') as fh:
            content = base64.b64encode(fh.read()).decode()
        sha = _gh('POST', f'{api}/git/blobs', token,
                  {'content': content, 'encoding': 'base64'})['sha']
        tree_items.append({'path': repo_path, 'mode': '100644', 'type': 'blob', 'sha': sha})
        print(f'        uploaded {repo_path}')

    tree = _gh('POST', f'{api}/git/trees', token,
               {'base_tree': base_tree, 'tree': tree_items})['sha']
    commit = _gh('POST', f'{api}/git/commits', token, {
        'message': f"Screener data as of {stamp['data']}", 'tree': tree,
        'parents': [parent]})['sha']
    _gh('PATCH', f'{api}/git/refs/heads/{GITHUB_BRANCH}', token, {'sha': commit})

    owner, repo = GITHUB_REPO.split('/')
    print(f'      committed {commit[:7]} - live in a minute or two at '
          f'https://{owner}.github.io/{repo}/')
    if GITHUB_XLSX:
        print(f'      workbook -> https://github.com/{GITHUB_REPO}/raw/{GITHUB_BRANCH}/{GITHUB_XLSX}')
    print('      if the page looks unchanged, hard refresh with Ctrl+Shift+R')


# ------------------------------------------------------------------ main -------------
def main():
    ap = argparse.ArgumentParser(description='Marquee portfolio and GICS sector screener builder')
    ap.add_argument('--data-dir', default=DATA_DIR, help='folder with the two CSVs')
    ap.add_argument('--out-dir', default=OUT_DIR, help='where to write the workbook and HTML')
    ap.add_argument('--no-html', action='store_true', help='skip the interactive HTML')
    ap.add_argument('--publish-dir', default=None,
                    help='also write index.html here, e.g. a checked-out website repo')
    ap.add_argument('--scrape', choices=['never', 'auto', 'always'], default=None,
                    help='override SCRAPE_DATAROMA')
    ap.add_argument('--no-upload', action='store_true',
                    help='skip the GitHub API upload, e.g. when a CI job commits instead')
    ap.add_argument('--stable-names', action='store_true',
                    help='drop the date from the output filenames')
    ap.add_argument('--web', action='store_true',
                    help='write a light index.html plus data.json instead of one big file')

    # Jupyter, Spyder and VS Code notebooks inject their own arguments such as
    # -f kernel-xxxx.json, which argparse would reject. Inside a notebook the command
    # line is ignored entirely and the DATA_DIR / OUT_DIR values above are used.
    in_notebook = ('ipykernel' in sys.modules or 'spyder_kernels' in sys.modules
                   or any(a.endswith('.json') for a in sys.argv[1:]))
    args = ap.parse_args([]) if in_notebook else ap.parse_known_args()[0]

    # command line beats the constants at the top, which is how the CI workflow drives
    # this script without editing the file
    global SCRAPE_DATAROMA, PUBLISH_DIR, GITHUB_UPLOAD, DATE_IN_FILENAME
    if args.scrape:
        SCRAPE_DATAROMA = args.scrape
    if args.publish_dir is not None:
        PUBLISH_DIR = args.publish_dir
    if args.stable_names:
        DATE_IN_FILENAME = False
    if args.no_upload or os.environ.get('GITHUB_ACTIONS') == 'true':
        # inside GitHub Actions the workflow commits the files itself, so the API
        # uploader is never needed and a personal token is never required
        GITHUB_UPLOAD = False
    if in_notebook:
        print('[note] notebook detected, command line ignored - using DATA_DIR and OUT_DIR '
              'from the top of the script')

    os.makedirs(args.out_dir, exist_ok=True)
    print(f'[1/6] data folder : {os.path.abspath(args.data_dir)}')
    if not os.path.isdir(args.data_dir):
        sys.exit(f'\n[ERROR] The folder does not exist:\n        '
                 f'{os.path.abspath(args.data_dir)}\n        Fix: correct DATA_DIR.\n')
    do_scrape, why = dataroma_needs_scrape(args.data_dir)
    if do_scrape:
        print(f'      scraping dataroma.com ({why})')
        have_old = any(classify(x) == 'marquee'
                       for x in glob.glob(os.path.join(args.data_dir, '*.csv')))
        try:
            scrape_dataroma(args.data_dir)
        except SystemExit as e:
            if not have_old:
                raise
            print(f'      [WARN] scrape failed, carrying on with the existing CSV')
            print(f'             reason: {str(e).strip().splitlines()[-1] if str(e) else e}')
    elif why:
        print(f'      dataroma scrape skipped - {why}')
    p_all, p_dr = find_files(args.data_dir)
    d_all, d_dr = file_date(p_all), file_date(p_dr)
    stamp = {'data': pretty_date(d_all), 'marquee': pretty_date(d_dr),
             'iso': d_all.isoformat(),
             'built': pretty_date(datetime.date.today())}
    print(f'      data as of  : {stamp["data"]}   (marquee list {stamp["marquee"]})')
    print(f'      universe    : {os.path.basename(p_all)}')
    print(f'      marquee     : {os.path.basename(p_dr)}')
    print(f'      output to   : {os.path.abspath(args.out_dir)}')

    a = load_universe(p_all)
    d = load_marquee(p_dr)
    print(f'[2/6] loaded {len(a)} universe rows, {len(d)} unique marquee rows')

    merged = match_marquee(a, d)
    matched = merged[merged['Matched Symbol'].notna()].copy()
    unid = merged[merged['Matched Symbol'].isna()].copy()
    matched['GICS Sector'] = assign_gics(pd.DataFrame({
        'Sector': matched['A_Sector'], 'Industry': matched['A_Industry']}))
    matched = matched.sort_values('Portfolio %', ascending=False).reset_index(drop=True)
    unid = unid.sort_values('Portfolio %', ascending=False).reset_index(drop=True)
    print(f'[3/6] matched {len(matched)}, unidentified {len(unid)}')

    wb = Workbook()
    sheet_readme(wb, len(matched), len(unid), len(a), {}, stamp)
    _, pos = sheet_marquee(wb, matched)
    sheet_unidentified(wb, unid)
    sheet_sector_summary(wb, matched, pos)
    a2 = sheet_all_stocks(wb, a, matched)

    # marquee ownership carried onto the screener rows
    wt = matched.set_index('Matched Symbol')['Portfolio %'].to_dict()
    cnt = matched.set_index('Matched Symbol')['Ownership Count'].to_dict()
    a2['Marquee Wt %'] = a2['Symbol'].map(wt)
    a2['Investors Holding'] = a2['Symbol'].map(cnt)

    pool = a2[a2['Marquee Flag'] == 'Marquee holding'] if SCREEN_UNIVERSE == 'marquee' else a2
    label = 'marquee holdings only' if SCREEN_UNIVERSE == 'marquee' else 'full universe'
    unmapped = int(pool['GICS Sector'].isna().sum())
    print(f"[4/6] screening {len(pool)} rows ({label}), THRESHOLD_MODE = '{THRESHOLD_MODE}'")
    if unmapped:
        print(f'      note: {unmapped} rows carry a TradingView sector with no GICS equivalent '
              f'(Miscellaneous / Government). They can pass All Sectors but reach no sector sheet.')

    def rule_line(applied):
        return '   |   '.join(f'{lbl} {op} {val}' for lbl, op, val in applied)

    counts = {}
    screened_all = set()
    for gics in screen_order():
        if gics == ALL_LABEL and ALL_SECTORS_USES_SECTOR_RULES:
            # each row judged on its own sector's block, then combined
            parts = []
            for g2 in SECTOR_MAP:
                p2, _, _ = screen(pool[pool['GICS Sector'] == g2], g2)
                parts.append(p2)
            unmapped = pool[pool['GICS Sector'].isna()]
            p3, applied, _ = screen(unmapped, ALL_LABEL)
            parts.append(p3)
            passed = pd.concat(parts) if parts else pool.iloc[0:0]
            sub = pool
            rule_text = ('each row judged against its own GICS sector block, then combined '
                         '(so this sheet is the union of the eleven sector sheets). Rows with no '
                         "GICS sector use the 'All Sectors' block: " + rule_line(applied))
        else:
            sub = pool if gics == ALL_LABEL else pool[pool['GICS Sector'] == gics]
            passed, applied, _ = screen(sub, gics)
            rule_text = rule_line(applied)
        counts[gics] = len(passed)
        screened_all.update(passed['Symbol'])
        sheet_sector_screen(wb, gics, passed, rule_text, label, stamp)
        us = int((passed['Country or region of registration'] == 'United States').sum())
        mq = int((passed['Marquee Flag'] == 'Marquee holding').sum())
        flag = '   <-- nothing passed, loosen the thresholds' if len(passed) == 0 else ''
        note = ''
        if gics == ALL_LABEL and ALL_SECTORS_USES_SECTOR_RULES:
            note = '   = union of the 11 sector screens + unmapped rows'
        print(f'      {gics:<24} pool {len(sub):>5}   pass {len(passed):>4} '
              f'(US {us}, non-US {len(passed) - us} | marquee {mq}, other {len(passed) - mq})'
              f'{flag}{note}')
    sheet_screen_settings(wb, counts, stamp)
    write_deep_feed(a2, screened_all, args.out_dir, PUBLISH_DIR, stamp)

    # refresh the Read Me pass-count line now that screens have run
    ws = wb['Read Me']
    for row in ws.iter_rows(min_col=2, max_col=3):
        if row[0].value == 'Passing counts this run':
            row[1].value = ', '.join(f'{g} {n}' for g, n in counts.items())

    wb.properties.title = f"Marquee Investor Portfolio and GICS Sector Screener - {stamp['data']}"
    wb.properties.creator = 'marquee_screener.py'
    suffix = f"_{stamp['iso']}" if DATE_IN_FILENAME else ''
    xlsx = os.path.join(args.out_dir, f'Marquee_Investor_Portfolio_Consolidated{suffix}.xlsx')
    wb.save(xlsx)
    print(f'[5/6] workbook saved -> {xlsx}')

    if not args.no_html:
        web = WEB_SPLIT or args.web
        html = os.path.join(args.out_dir,
                            'index.html' if web else f'Sector_Screener{suffix}.html')
        n_html, n_data = build_html(a2, html, stamp, split=web)
        print(f'[6/6] screener saved -> {html}  ({n_html/1024:,.0f} KB)')
        if n_data:
            print(f'      data file      -> {os.path.join(args.out_dir, "data.json")}  '
                  f'({n_data/1024:,.0f} KB)   keep both files together')
        if PUBLISH_DIR:
            if not os.path.isdir(PUBLISH_DIR):
                print(f'      [WARN] PUBLISH_DIR does not exist, skipped: {PUBLISH_DIR}')
            else:
                import shutil
                dest = os.path.join(PUBLISH_DIR, 'index.html')
                shutil.copyfile(html, dest)
                print(f'      published copy -> {dest}')
                if n_data:
                    shutil.copyfile(os.path.join(args.out_dir, 'data.json'),
                                    os.path.join(PUBLISH_DIR, 'data.json'))
                    print(f'      published copy -> {os.path.join(PUBLISH_DIR, "data.json")}')
                print('      next: cd into that folder, then')
                print(f'        git add index.html && git commit -m "screener {stamp["iso"]}" && git push')
        if GITHUB_UPLOAD:
            payload = [(html, GITHUB_PATH), (xlsx, GITHUB_XLSX)]
            payload += [(os.path.join(args.out_dir, f), f) for f in GITHUB_FEED]
            publish_to_github(payload, stamp)
    else:
        print('[6/6] HTML skipped')
    print('\nDone. The "Screen Settings" sheet lists the final thresholds for every sector.')


if __name__ == '__main__':
    main()
